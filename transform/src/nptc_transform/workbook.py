"""Reads the SPIA workbook cell-by-cell, preserving the original cell type.

FR-63 documents the published column layout; ``COLUMN_ROLES`` maps that layout
to a small set of roles so the defect scanner (``cell_defects.py``) can find
the code column without hard-coding a column letter. Cell *type* is captured
as read - never coerced - because the type itself is often the defect: a code
cell typed as a number rather than text is PRD Appendix A.2's finding, and
type coercion here would make that defect unrepresentable rather than surface
it (FR-06, FR-70's "MUST NOT silently repair").
"""

from __future__ import annotations

import datetime
import re
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterator
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path
from typing import Protocol

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from nptc_shared.text import escape_invisible
from nptc_transform.cellref import CellRef


class _RawCell(Protocol):
    """The subset of openpyxl's cell interface used here.

    ``value`` is a read-only property, not a plain attribute: openpyxl's
    ``MergedCell.value`` is typed as the literal ``None`` rather than
    ``object``, and Protocol attribute matching requires exact (invariant)
    types for a mutable attribute but only covariant matching for a
    property - a plain ``value: object`` attribute would reject
    ``MergedCell``.
    """

    data_type: str

    @property
    def value(self) -> object: ...


class WorkbookReadError(Exception):
    """The workbook could not be opened or read.

    Wraps whatever openpyxl or the zip layer raises for a missing, corrupt or
    non-``.xlsx`` file, so the CLI can turn it into a plain usage error
    (exit 2) instead of a traceback.
    """


class CellType(StrEnum):
    """The cell's original storage type, exactly as recorded in the workbook."""

    TEXT = "text"
    NUMBER = "number"
    DATE = "date"
    BOOLEAN = "boolean"
    FORMULA = "formula"
    ERROR = "error"


class ColumnRole(StrEnum):
    """The semantic role of a column, identified by its header text."""

    CODE = "code"
    PREFERRED_TERM = "preferred_term"
    SYNONYMS = "synonyms"
    FSN = "fsn"
    GUIDANCE = "guidance"
    LENGTH = "length"
    DISCIPLINE = "discipline"
    SUBGROUP = "subgroup"
    SPECIMEN = "specimen"
    VERSION = "version"
    HISTORY = "history"
    UNKNOWN = "unknown"


def _normalise_header(header: str) -> str:
    """Normalises a header for lookup in ``COLUMN_ROLES``.

    Strips, casefolds and collapses internal whitespace runs (including
    non-ASCII whitespace such as U+00A0) to a single space, so a header that
    itself carries an Appendix A.1/A.3 defect still resolves to its role.
    """
    collapsed = re.sub(r"\s+", " ", header, flags=re.UNICODE)
    return collapsed.strip().casefold()


# FR-63's documented published layout.
COLUMN_ROLES: dict[str, ColumnRole] = {
    "rcpa preferred term": ColumnRole.PREFERRED_TERM,
    "rcpa synonyms": ColumnRole.SYNONYMS,
    "usage guidance": ColumnRole.GUIDANCE,
    "length": ColumnRole.LENGTH,
    "discipline": ColumnRole.DISCIPLINE,
    "subgroup": ColumnRole.SUBGROUP,
    "specimen": ColumnRole.SPECIMEN,
    "terminology binding (snomed ct-au)": ColumnRole.CODE,
    "snomed ct fully specified name": ColumnRole.FSN,
    "version": ColumnRole.VERSION,
    "history": ColumnRole.HISTORY,
}


def column_role(header: str) -> ColumnRole:
    """Looks up the role of a column by its header text (see ``COLUMN_ROLES``)."""
    return COLUMN_ROLES.get(_normalise_header(header), ColumnRole.UNKNOWN)


@dataclass(frozen=True)
class Cell:
    """One cell, with its position, header, role, original type and text.

    ``raw`` is the value openpyxl handed back (``int``, ``float``, ``bool``,
    ``datetime``, ``str`` or ``None``) - never re-interpreted. ``text`` is a
    verbatim rendering for display and scanning; see ``_render_text``.
    """

    sheet: str
    row: int
    column: int
    column_letter: str
    header: str
    role: ColumnRole
    cell_type: CellType
    text: str
    raw: object

    @property
    def reference(self) -> CellRef:
        return CellRef(sheet=self.sheet, column_letter=self.column_letter, row=self.row)


@dataclass(frozen=True)
class Sheet:
    """One worksheet's headers and non-empty cells."""

    name: str
    headers: tuple[str, ...]
    cells: tuple[Cell, ...]


_DATA_TYPE_TO_CELL_TYPE: dict[str, CellType] = {
    "s": CellType.TEXT,
    "inlineStr": CellType.TEXT,
    "str": CellType.TEXT,
    "n": CellType.NUMBER,
    "d": CellType.DATE,
    "b": CellType.BOOLEAN,
    "f": CellType.FORMULA,
    "e": CellType.ERROR,
}


def _cell_type(data_type: str) -> CellType:
    return _DATA_TYPE_TO_CELL_TYPE.get(data_type, CellType.TEXT)


def _render_text(value: object) -> str:
    """Renders ``value`` to text without ever passing an int through float.

    Order matters: ``bool`` is checked before ``int`` because ``bool`` is an
    ``int`` subclass in Python. A code short enough to survive as an exact
    ``int`` (up to roughly 4.5e15) renders to its exact digits (FR-06) rather
    than being routed through ``float`` and losing them. A code long enough
    that Excel itself already stores it as a float - which is precisely the
    corruption Appendix A.2 is about - renders via ``repr()``; that value is
    already lossy by the time it reaches this function, and no rendering
    choice here recovers it.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, datetime.datetime | datetime.date | datetime.time):
        return value.isoformat()
    if isinstance(value, float):
        return repr(value)
    return str(value)


def _iter_sheet_cells(
    sheet_name: str, headers: tuple[str, ...], rows: Iterator[tuple[_RawCell, ...]]
) -> Iterator[Cell]:
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, raw_cell in enumerate(row, start=1):
            value = raw_cell.value
            if value is None:
                continue
            header = headers[col_idx - 1] if col_idx <= len(headers) else ""
            yield Cell(
                sheet=sheet_name,
                row=row_idx,
                column=col_idx,
                column_letter=get_column_letter(col_idx),
                header=header,
                role=column_role(header),
                cell_type=_cell_type(raw_cell.data_type),
                text=_render_text(value),
                raw=value,
            )


def read_workbook(path: Path) -> tuple[Sheet, ...]:
    """Reads every worksheet in ``path``, row 1 as headers, empty cells skipped.

    Opens ``read_only=True`` (the workbook is never mutated) and
    ``data_only=False`` - a formula cell's original type must be captured as
    a formula, not the value Excel last cached for it.
    """
    workbook = None
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
        sheets = []
        for worksheet in workbook.worksheets:
            # A worksheet title can itself carry an Appendix A.1 defect (Excel
            # permits U+00A0 in a sheet name) - escaped once here so it can
            # never put a raw invisible character into a Cell.reference, and
            # therefore into a report, via the sheet side of the reference.
            sheet_name = escape_invisible(worksheet.title)
            row_iter = worksheet.iter_rows()
            try:
                header_row = next(row_iter)
            except StopIteration:
                sheets.append(Sheet(name=sheet_name, headers=(), cells=()))
                continue
            headers = tuple(_render_text(cell.value) for cell in header_row)
            cells = tuple(_iter_sheet_cells(sheet_name, headers, row_iter))
            sheets.append(Sheet(name=sheet_name, headers=headers, cells=cells))
        return tuple(sheets)
    except (
        InvalidFileException,
        zipfile.BadZipFile,
        ET.ParseError,
        KeyError,
        IndexError,
        ValueError,
        OSError,
    ) as exc:
        # openpyxl's read-only mode parses a worksheet's dimensions as soon as
        # it opens (eagerly, inside load_workbook), but the rest of the sheet
        # lazily, as it's iterated - so a corrupt or unreadable workbook can
        # fail at either point, and both need the same treatment here.
        # ValueError/IndexError cover content-level corruption openpyxl makes
        # no promise not to raise: parse_cell() calls int()/float() on a
        # numeric cell's raw text (ValueError on garbage), from_ISO8601() on a
        # malformed date (ValueError), and indexes into sharedStrings by
        # position for a text cell (IndexError on a truncated sharedStrings
        # part) - openpyxl/worksheet/_reader.py, no try/except around any of
        # them.
        raise WorkbookReadError(f"could not read workbook {path}: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()
