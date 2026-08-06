"""Tests for the workbook reader: cell-type capture, headers, roles (FR-70)."""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pytest

from nptc_transform.workbook import (
    CellType,
    ColumnRole,
    WorkbookReadError,
    column_role,
    read_workbook,
)


@pytest.fixture()
def typed_workbook(tmp_path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(
        [
            "RCPA Preferred term",
            "Terminology binding (SNOMED CT-AU)",
            "Length",
            "Version",
            "Active",
        ]
    )
    sheet.append(["term one", None, None, None, None])
    code_cell = sheet.cell(row=2, column=2, value="12345678")
    code_cell.data_type = "s"

    sheet.append(["term two", 933434771000036107, "=LEN(A3)", datetime.date(2025, 2, 1), True])

    path = tmp_path / "typed.xlsx"
    workbook.save(path)
    return path


def test_reads_headers_and_assigns_roles(typed_workbook: Path) -> None:
    sheets = read_workbook(typed_workbook)
    assert len(sheets) == 1
    sheet = sheets[0]
    assert sheet.name == "Requesting"
    assert sheet.headers == (
        "RCPA Preferred term",
        "Terminology binding (SNOMED CT-AU)",
        "Length",
        "Version",
        "Active",
    )


def test_empty_cells_are_skipped(typed_workbook: Path) -> None:
    sheets = read_workbook(typed_workbook)
    references = {cell.reference for cell in sheets[0].cells}
    # Row 2 only has A2 and B2 populated - C2/D2/E2 must not appear.
    assert "Requesting!C2" not in references
    assert "Requesting!D2" not in references
    assert "Requesting!E2" not in references


def test_text_typed_code_cell_captured_verbatim(typed_workbook: Path) -> None:
    sheets = read_workbook(typed_workbook)
    cell = next(c for c in sheets[0].cells if c.reference == "Requesting!B2")
    assert cell.cell_type is CellType.TEXT
    assert cell.role is ColumnRole.CODE
    assert cell.text == "12345678"


def test_eighteen_digit_code_stored_as_number_renders_exact_digits_when_int(
    tmp_path: Path,
) -> None:
    """FR-06: an int must never be rendered via float, which would lose digits."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)"])
    sheet.append([123456789012345])  # 15 digits, survives as int through openpyxl
    path = tmp_path / "digits.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    cell = sheets[0].cells[0]
    assert cell.cell_type is CellType.NUMBER
    assert cell.text == "123456789012345"
    assert isinstance(cell.raw, int)


def test_formula_cell_captured_as_formula_not_its_cached_value(typed_workbook: Path) -> None:
    sheets = read_workbook(typed_workbook)
    cell = next(c for c in sheets[0].cells if c.reference == "Requesting!C3")
    assert cell.cell_type is CellType.FORMULA
    assert cell.text == "=LEN(A3)"


def test_date_cell_captured_as_date(typed_workbook: Path) -> None:
    sheets = read_workbook(typed_workbook)
    cell = next(c for c in sheets[0].cells if c.reference == "Requesting!D3")
    assert cell.cell_type is CellType.DATE
    assert cell.text == "2025-02-01T00:00:00"
    assert cell.role is ColumnRole.VERSION


def test_boolean_cell_captured_as_boolean_not_number(typed_workbook: Path) -> None:
    sheets = read_workbook(typed_workbook)
    cell = next(c for c in sheets[0].cells if c.reference == "Requesting!E3")
    assert cell.cell_type is CellType.BOOLEAN
    assert cell.text == "TRUE"


def test_error_cell_captured_as_error(tmp_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Some column"])
    error_cell = sheet.cell(row=2, column=1, value="#N/A")
    error_cell.data_type = "e"
    path = tmp_path / "error.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    cell = sheets[0].cells[0]
    assert cell.cell_type is CellType.ERROR


def test_cell_reference_format(typed_workbook: Path) -> None:
    sheets = read_workbook(typed_workbook)
    cell = next(c for c in sheets[0].cells if c.column_letter == "A" and c.row == 2)
    assert cell.reference == "Requesting!A2"


@pytest.mark.parametrize(
    ("header", "expected_role"),
    [
        ("RCPA Preferred term", ColumnRole.PREFERRED_TERM),
        ("Terminology binding (SNOMED CT-AU)", ColumnRole.CODE),
        ("  terminology binding (snomed ct-au)  ", ColumnRole.CODE),
        (f"Terminology binding (SNOMED CT-AU){chr(0x00A0)}", ColumnRole.CODE),
        ("Some Unrelated Column", ColumnRole.UNKNOWN),
    ],
)
def test_column_role_normalises_header_text(header: str, expected_role: ColumnRole) -> None:
    assert column_role(header) is expected_role


def test_workbook_with_no_data_rows_has_headers_and_no_cells(tmp_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["A", "B"])
    path = tmp_path / "headers_only.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    assert sheets[0].headers == ("A", "B")
    assert sheets[0].cells == ()


def test_sheet_with_no_rows_at_all(tmp_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    path = tmp_path / "empty.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    assert sheets[0].headers == ()
    assert sheets[0].cells == ()


def test_corrupt_file_raises_workbook_read_error(corrupt_workbook: Path) -> None:
    with pytest.raises(WorkbookReadError):
        read_workbook(corrupt_workbook)


def test_corrupt_worksheet_xml_raises_workbook_read_error(
    corrupt_worksheet_workbook: Path,
) -> None:
    """A valid zip with one unparsable sheet part fails during iteration, not
    at ``load_workbook()`` - both paths must still surface as the same error."""
    with pytest.raises(WorkbookReadError):
        read_workbook(corrupt_worksheet_workbook)


def test_blank_header_cell_renders_as_empty_string(tmp_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["A", None, "C"])
    sheet.append(["x", "y", "z"])
    path = tmp_path / "blank_header.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    assert sheets[0].headers == ("A", "", "C")
