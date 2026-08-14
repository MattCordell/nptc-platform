"""Report-only is the default mode; --emit-dataset refuses cleanly (FR-70)."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest
from click.testing import Result
from typer.testing import CliRunner

from nptc_shared.terminology.errors import TerminologyTransportError
from nptc_shared.terminology.models import PROCEDURE_ROOT_CODE, Operation
from nptc_shared.terminology.stub import StubConcept, StubTerminologyClient
from nptc_transform import __version__
from nptc_transform.cli import app

runner = CliRunner()

CLEAN_CODE = "122192001"


def _tree(root: Path) -> set[str]:
    return {str(p.relative_to(root)) for p in root.rglob("*")}


class _ContextStub(StubTerminologyClient):
    """The stub, wearing ``OntoserverClient``'s context-manager shape so it
    can stand in for it where the CLI builds one."""

    def __enter__(self) -> _ContextStub:
        return self

    def __exit__(self, *_exc_info: object) -> None:
        return None


@pytest.fixture()
def clean_bindings_workbook(tmp_path: Path) -> Path:
    """One row, one valid code, no cell defects - so the exit code reflects
    the terminology pass and nothing else."""
    path = tmp_path / "clean.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.cell(row=2, column=1, value="Acanthamoeba culture")
    code_cell = sheet.cell(row=2, column=2, value=CLEAN_CODE)
    code_cell.data_type = "s"
    workbook.save(path)
    return path


def _install_stub(monkeypatch: pytest.MonkeyPatch, stub: StubTerminologyClient) -> None:
    monkeypatch.setattr("nptc_transform.cli.OntoserverClient", lambda _config: stub)


def test_cli_version_command_runs() -> None:
    result = runner.invoke(app, ["version"])
    assert result.exit_code == 0
    assert result.stdout.strip() == __version__


@pytest.mark.req("FR-70")
def test_report_only_is_the_default_and_writes_only_the_report_dir(
    tmp_path: Path, sample_workbook: Path
) -> None:
    report_dir = tmp_path / "report"
    before = _tree(tmp_path)

    result = runner.invoke(
        app, ["run", "--workbook", str(sample_workbook), "--report-dir", str(report_dir)]
    )

    assert result.exit_code == 0, result.output
    assert (report_dir / "report.json").is_file()
    assert (report_dir / "report.md").is_file()

    # Exactly the report dir and its two files - not "something starting with
    # report", which would pass for a stray sibling too.
    assert _tree(tmp_path) - before == {
        str(Path("report")),
        str(Path("report/report.json")),
        str(Path("report/report.md")),
    }
    # The default --report-dir is relative, so a regression that ignores the
    # option writes into the process CWD, which is outside tmp_path entirely.
    assert not (Path.cwd() / "transform-report").exists()


@pytest.mark.req("FR-76")
def test_emit_dataset_without_release_name_is_a_usage_error_and_writes_nothing(
    tmp_path: Path, sample_workbook: Path
) -> None:
    report_dir = tmp_path / "report"
    before = _tree(tmp_path)

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
        ],
    )

    assert result.exit_code == 2
    assert "--release-name" in result.output
    assert _tree(tmp_path) == before
    assert not report_dir.exists()


@pytest.mark.req("FR-76")
def test_emit_dataset_with_a_malformed_release_name_is_a_usage_error(
    tmp_path: Path, sample_workbook: Path
) -> None:
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026",
        ],
    )

    assert result.exit_code == 2
    assert "--release-name" in result.output
    assert not report_dir.exists()


@pytest.mark.req("FR-57")
def test_emit_dataset_with_an_impossible_month_is_a_usage_error(
    tmp_path: Path, sample_workbook: Path
) -> None:
    """The value lands verbatim in baseline_release.name, which FR-60 will
    later diff against - an impossible month like '13' must be refused here,
    not accepted and only discovered downstream."""
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-13",
        ],
    )

    assert result.exit_code == 2
    assert "--release-name" in result.output
    assert not report_dir.exists()


@pytest.mark.req("FR-76")
def test_release_name_without_emit_dataset_is_a_usage_error(
    tmp_path: Path, sample_workbook: Path
) -> None:
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--release-name",
            "2026-06",
        ],
    )

    assert result.exit_code == 2
    assert "--emit-dataset" in result.output
    assert not report_dir.exists()


@pytest.mark.req("FR-76")
def test_emit_dataset_writes_the_report_and_the_import_dataset(
    tmp_path: Path, sample_workbook: Path
) -> None:
    report_dir = tmp_path / "report"
    before = _tree(tmp_path)

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-06",
        ],
    )

    assert result.exit_code == 0, result.output
    assert (report_dir / "report.json").is_file()
    assert (report_dir / "report.md").is_file()
    assert (report_dir / "import-dataset.json").is_file()
    assert _tree(tmp_path) - before == {
        str(Path("report")),
        str(Path("report/report.json")),
        str(Path("report/report.md")),
        str(Path("report/import-dataset.json")),
    }

    payload = json.loads((report_dir / "import-dataset.json").read_text(encoding="utf-8"))
    assert payload["baseline_release"] == {
        "name": "2026-06",
        "note": "Synthetic baseline release representing the state at seeding (FR-76).",
    }
    assert len(payload["entries"]) == 1
    entry = payload["entries"][0]
    assert entry["business_key"] == "NPTC-000001"
    assert entry["code_bindings"][0]["code"] == "10000006"
    assert isinstance(entry["code_bindings"][0]["code"], str)


@pytest.mark.req("FR-71")
def test_emit_dataset_blocks_on_a_blocking_finding_and_writes_no_dataset(
    tmp_path: Path, annex_a_workbook: Path
) -> None:
    """A blocking finding aborts emission: exit 1, report written, no
    dataset file (PRD:310's "the seeded baseline cannot be created until
    RCPA-QAP resolves those collisions editorially")."""
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(annex_a_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-06",
        ],
    )

    assert result.exit_code == 1, result.output
    assert (report_dir / "report.json").is_file()
    assert not (report_dir / "import-dataset.json").exists()


def _run_emit_dataset(tmp_path: Path, workbook_path: Path) -> tuple[Path, Result]:
    report_dir = tmp_path / "report"
    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(workbook_path),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-06",
        ],
    )
    return report_dir, result


def _workbook_with_code(tmp_path: Path, name: str, code: str) -> Path:
    path = tmp_path / name
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.cell(row=2, column=1, value="A term")
    code_cell = sheet.cell(row=2, column=2, value=code)
    code_cell.data_type = "s"
    workbook.save(path)
    return path


@pytest.mark.req("FR-06")
def test_emit_dataset_alone_blocks_on_a_malformed_code_with_no_check_terminology(
    tmp_path: Path,
) -> None:
    """Issue #131: ``--emit-dataset`` without ``--check-terminology`` must
    not seed a code that isn't even a well-formed SCTID - well-formedness is
    an offline, free check and must never be gated behind the opt-in
    network validation pass."""
    workbook_path = _workbook_with_code(tmp_path, "not_a_code.xlsx", "not-a-code")

    report_dir, result = _run_emit_dataset(tmp_path, workbook_path)

    assert result.exit_code == 1, result.output
    assert "CODE_NOT_WELL_FORMED" in (report_dir / "report.json").read_text(encoding="utf-8")
    assert not (report_dir / "import-dataset.json").exists()


@pytest.mark.req("FR-06")
def test_emit_dataset_alone_blocks_on_a_bad_check_digit_with_no_check_terminology(
    tmp_path: Path,
) -> None:
    """Same as above, but for a code that has the right shape (6-18 digits)
    and still fails the Verhoeff check digit - shape alone isn't enough."""
    workbook_path = _workbook_with_code(tmp_path, "bad_check_digit.xlsx", "123456789")

    report_dir, result = _run_emit_dataset(tmp_path, workbook_path)

    assert result.exit_code == 1, result.output
    assert "CODE_NOT_WELL_FORMED" in (report_dir / "report.json").read_text(encoding="utf-8")
    assert not (report_dir / "import-dataset.json").exists()


@pytest.mark.req("FR-06")
def test_emit_dataset_alone_blocks_on_an_nbsp_that_corrects_to_an_interior_space(
    tmp_path: Path,
) -> None:
    """An interior U+00A0 is corrected to an ordinary space before this
    check runs (the same correction ``dataset.py`` applies before seeding),
    so a code with an interior NBSP between two digit groups must be caught
    once corrected - not judged well-formed because the raw, uncorrected
    cell text happened to be all digits plus one invisible character."""
    nbsp = chr(0x00A0)
    workbook_path = _workbook_with_code(tmp_path, "nbsp_code.xlsx", f"123{nbsp}456")

    report_dir, result = _run_emit_dataset(tmp_path, workbook_path)

    assert result.exit_code == 1, result.output
    assert "CODE_NOT_WELL_FORMED" in (report_dir / "report.json").read_text(encoding="utf-8")
    assert not (report_dir / "import-dataset.json").exists()


@pytest.mark.req("FR-06")
def test_emit_dataset_alone_still_emits_cleanly_for_a_well_formed_code(
    tmp_path: Path,
) -> None:
    """The companion positive case: a well-formed, Verhoeff-valid code still
    emits with exit 0 and no new findings - this check must not flag codes
    it has no business flagging."""
    workbook_path = _workbook_with_code(tmp_path, "clean_code.xlsx", CLEAN_CODE)

    report_dir, result = _run_emit_dataset(tmp_path, workbook_path)

    assert result.exit_code == 0, result.output
    assert "CODE_NOT_WELL_FORMED" not in (report_dir / "report.json").read_text(encoding="utf-8")
    payload = json.loads((report_dir / "import-dataset.json").read_text(encoding="utf-8"))
    assert payload["entries"][0]["code_bindings"][0]["code"] == CLEAN_CODE


@pytest.mark.req("FR-71")
def test_a_blocking_run_removes_a_stale_dataset_from_an_earlier_successful_run(
    tmp_path: Path, sample_workbook: Path, annex_a_workbook: Path
) -> None:
    """A prior clean run leaves ``import-dataset.json`` next to its report.
    A later run into the same ``--report-dir`` that hits a blocking finding
    must not leave that dataset behind: the refreshed report says the import
    is blocked, so a dataset from a previous, unrelated run sitting beside it
    is a stale-and-misleading artifact, not a still-valid one (issue #130)."""
    report_dir = tmp_path / "report"

    first = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-06",
        ],
    )
    assert first.exit_code == 0, first.output
    assert (report_dir / "import-dataset.json").is_file()

    second = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(annex_a_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-06",
        ],
    )

    assert second.exit_code == 1, second.output
    assert (report_dir / "report.json").is_file()
    assert not (report_dir / "import-dataset.json").exists()


@pytest.mark.req("FR-71")
def test_a_report_only_run_removes_a_stale_dataset_from_an_earlier_successful_run(
    tmp_path: Path, sample_workbook: Path
) -> None:
    """The same stale-dataset invariant as the blocked case, reached via the
    success path instead: a run into the same ``--report-dir`` that omits
    ``--emit-dataset`` refreshes the report but must not leave a previous
    run's ``import-dataset.json`` sitting beside it either (issue #130)."""
    report_dir = tmp_path / "report"

    first = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-06",
        ],
    )
    assert first.exit_code == 0, first.output
    assert (report_dir / "import-dataset.json").is_file()

    second = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
        ],
    )

    assert second.exit_code == 0, second.output
    assert (report_dir / "report.json").is_file()
    assert not (report_dir / "import-dataset.json").exists()


@pytest.mark.req("FR-70")
def test_emit_dataset_blocks_and_writes_no_dataset_for_a_row_missing_its_preferred_term(
    tmp_path: Path,
) -> None:
    """A row that resolves a code binding but carries no 'RCPA Preferred
    term' value must block emission, exactly like any other data defect -
    silently omitting the row from import-dataset.json (as build_dataset did
    before MISSING_PREFERRED_TERM existed) is the P0-9/#31 hazard this
    regression guards against."""
    workbook_path = tmp_path / "missing_preferred_term.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.cell(row=2, column=2, value="12345678").data_type = "s"
    workbook.save(workbook_path)
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(workbook_path),
            "--report-dir",
            str(report_dir),
            "--emit-dataset",
            "--release-name",
            "2026-06",
        ],
    )

    assert result.exit_code == 1, result.output
    assert "MISSING_PREFERRED_TERM" in (report_dir / "report.json").read_text(encoding="utf-8")
    assert not (report_dir / "import-dataset.json").exists()


@pytest.mark.req("FR-70")
def test_report_only_can_be_stated_explicitly(tmp_path: Path, sample_workbook: Path) -> None:
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--report-only",
        ],
    )

    assert result.exit_code == 0, result.output
    assert (report_dir / "report.json").is_file()


@pytest.mark.req("FR-70")
def test_report_only_and_emit_dataset_together_are_a_usage_error(
    tmp_path: Path, sample_workbook: Path
) -> None:
    """Asking for both modes at once must be refused, not silently resolved."""
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(sample_workbook),
            "--report-dir",
            str(report_dir),
            "--report-only",
            "--emit-dataset",
        ],
    )

    assert result.exit_code == 2
    assert "mutually exclusive" in result.output
    assert not report_dir.exists()


@pytest.mark.req("FR-70")
def test_report_dir_pointing_at_a_file_is_a_usage_error(
    tmp_path: Path, sample_workbook: Path
) -> None:
    """An unwritable --report-dir must explain itself, not raise a traceback.

    Exit 1 is reserved for the report containing blocking findings (FR-71),
    so a filesystem refusal has to land on 2 with a message the operator can
    act on.
    """
    not_a_dir = tmp_path / "report"
    not_a_dir.write_text("", encoding="utf-8")

    result = runner.invoke(
        app, ["run", "--workbook", str(sample_workbook), "--report-dir", str(not_a_dir)]
    )

    assert result.exit_code == 2, result.output
    assert result.exception is None or isinstance(result.exception, SystemExit)
    assert "Traceback" not in result.output


@pytest.mark.req("FR-70")
def test_missing_workbook_is_a_usage_error(tmp_path: Path) -> None:
    result = runner.invoke(app, ["run", "--workbook", str(tmp_path / "does-not-exist.xlsx")])

    assert result.exit_code == 2


@pytest.mark.req("FR-70")
def test_unreadable_workbook_is_a_usage_error_not_a_traceback(
    tmp_path: Path, corrupt_workbook: Path
) -> None:
    """A file that exists and is readable but isn't a valid workbook (P0-2)
    must be refused the same way a missing file is - exit 2, no traceback."""
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app, ["run", "--workbook", str(corrupt_workbook), "--report-dir", str(report_dir)]
    )

    assert result.exit_code == 2, result.output
    assert "Traceback" not in result.output
    assert not report_dir.exists()
    # WorkbookReadError's own message already says "could not read workbook
    # ...": a regression that wraps it a second time doubles this phrase.
    assert result.output.count("could not read") == 1


@pytest.mark.req("FR-74")
def test_check_terminology_validates_the_bindings_and_records_the_run(
    tmp_path: Path, clean_bindings_workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    report_dir = tmp_path / "report"
    _install_stub(
        monkeypatch,
        _ContextStub(
            concepts=[
                StubConcept(
                    code=CLEAN_CODE,
                    fsn="Acanthamoeba culture (procedure)",
                    parents=(PROCEDURE_ROOT_CODE,),
                )
            ],
            resolved_version={"au": "http://snomed.info/sct/32506021000036107/version/20260531"},
        ),
    )

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(clean_bindings_workbook),
            "--report-dir",
            str(report_dir),
            "--check-terminology",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads((report_dir / "report.json").read_text(encoding="utf-8"))
    assert payload["terminology"]["codes_checked"] == 1
    assert payload["finding_count"] == 0
    # No "SNOMED CT Fully Specified Name" column on this fixture - every
    # checkable code cell is a row FR-97 could not reconcile, not zero rows.
    assert payload["designations"] == {
        "labels_reconciled": 0,
        "labels_not_reconciled": 1,
        "label_confirmations": 0,
    }


@pytest.mark.req("FR-79")
def test_check_terminology_run_reports_sweep_backed_misspellings(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """End to end through the CLI, not just via ``check_misspellings`` called
    directly (``test_misspelling.py``'s own helper reimplements
    ``pipeline.py``'s wiring rather than exercising it - see the review that
    flagged this gap). A sweep-backed run must record
    ``misspellings.authority_source == "SWEEP"`` in the report, and the
    Amylase/Amylose whitelist suppression must hold with the real CLI ->
    pipeline plumbing in between: a regression that silently downgrades
    ``pipeline.py``'s ``check_misspellings(sheets, results=outcome.results)``
    to the no-sweep ``check_misspellings(sheets)`` call must fail this test.
    """
    amylase_code, amylose_code = "700000010", "700000023"
    workbook_path = tmp_path / "amylase.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    rows: list[tuple[str, str | None]] = [
        ("Amylose", amylose_code),
        ("Amylase", amylase_code),
        ("Amylase panel", amylase_code),
        ("Amylase ratio", amylase_code),
    ]
    for index, (term, code) in enumerate(rows, start=2):
        sheet.cell(row=index, column=1, value=term)
        if code is not None:
            code_cell = sheet.cell(row=index, column=2, value=code)
            code_cell.data_type = "s"
    workbook.save(workbook_path)

    report_dir = tmp_path / "report"
    _install_stub(
        monkeypatch,
        _ContextStub(
            concepts=[
                StubConcept(
                    code=amylase_code,
                    fsn="Amylase (substance)",
                    parents=(PROCEDURE_ROOT_CODE,),
                ),
                StubConcept(
                    code=amylose_code,
                    fsn="Amylose (substance)",
                    parents=(PROCEDURE_ROOT_CODE,),
                ),
            ],
            resolved_version={"au": "http://snomed.info/sct/32506021000036107/version/20260531"},
        ),
    )

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(workbook_path),
            "--report-dir",
            str(report_dir),
            "--check-terminology",
        ],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads((report_dir / "report.json").read_text(encoding="utf-8"))
    assert payload["misspellings"]["authority_source"] == "SWEEP"
    assert not any(
        defect_class["code"] in ("PROBABLE_MISSPELLING", "INCONSISTENT_SPELLING")
        for defect_class in payload["defect_classes"]
    )


@pytest.mark.req("FR-97")
def test_check_terminology_blocks_on_a_designation_defect(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """End to end: a published label matching no designation on the bound
    concept (PRD row 22's own shape) aborts the import through the CLI."""
    workbook_path = tmp_path / "row22.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(
        [
            "RCPA Preferred term",
            "Terminology binding (SNOMED CT-AU)",
            "SNOMED CT Fully Specified Name",
        ]
    )
    sheet.cell(row=2, column=1, value="Acanthamoeba culture")
    code_cell = sheet.cell(row=2, column=2, value=CLEAN_CODE)
    code_cell.data_type = "s"
    sheet.cell(row=2, column=3, value="Acanthamoeba species culture")
    workbook.save(workbook_path)

    report_dir = tmp_path / "report"
    _install_stub(
        monkeypatch,
        _ContextStub(
            concepts=[
                StubConcept(
                    code=CLEAN_CODE,
                    fsn="Acanthamoeba culture (procedure)",
                    parents=(PROCEDURE_ROOT_CODE,),
                )
            ],
            resolved_version={"au": "http://snomed.info/sct/32506021000036107/version/20260531"},
        ),
    )

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(workbook_path),
            "--report-dir",
            str(report_dir),
            "--check-terminology",
        ],
    )

    assert result.exit_code == 1, result.output
    assert "import blocked" in result.output
    payload = json.loads((report_dir / "report.json").read_text(encoding="utf-8"))
    assert payload["designations"]["labels_reconciled"] == 1
    assert any(
        defect_class["code"] == "LABEL_MATCHES_NO_DESIGNATION"
        for defect_class in payload["defect_classes"]
    )


@pytest.mark.req("FR-70")
def test_a_run_without_check_terminology_records_that_no_sweep_ran(
    tmp_path: Path, clean_bindings_workbook: Path
) -> None:
    """The flag is opt-in: without it the run neither reads NPTC_TX_* nor
    reports the codes as validated."""
    report_dir = tmp_path / "report"

    result = runner.invoke(
        app,
        ["run", "--workbook", str(clean_bindings_workbook), "--report-dir", str(report_dir)],
    )

    assert result.exit_code == 0, result.output
    payload = json.loads((report_dir / "report.json").read_text(encoding="utf-8"))
    assert payload["terminology"] is None
    assert payload["designations"] is None


@pytest.mark.req("FR-52")
def test_a_malformed_nptc_tx_chunk_size_is_a_usage_error_not_a_terminology_failure(
    tmp_path: Path, clean_bindings_workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """``TerminologyConfigError`` is a ``TerminologyError`` subclass, but a
    deployment typo is a usage error (exit 2) an operator should fix, not a
    transient server failure (exit 3) that would get retried indefinitely
    instead. Raised by ``TerminologyConfig.from_env()`` before the
    ``OntoserverClient`` is ever constructed, so no stub is needed here."""
    report_dir = tmp_path / "report"
    monkeypatch.setenv("NPTC_TX_CHUNK_SIZE", "lots")

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(clean_bindings_workbook),
            "--report-dir",
            str(report_dir),
            "--check-terminology",
        ],
    )

    assert result.exit_code == 2, result.output
    assert not report_dir.exists()
    assert "invalid terminology configuration" in result.output
    assert "terminology validation failed" not in result.output


@pytest.mark.req("FR-70")
def test_a_corrupt_workbook_reports_its_own_error_even_with_a_bad_nptc_tx_value(
    tmp_path: Path, corrupt_workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The workbook is read before any terminology configuration is even
    touched - a simultaneously malformed ``NPTC_TX_*`` value must not
    pre-empt the clearer, actionable workbook error with a config-specific
    one the operator would have to untangle from the real problem."""
    report_dir = tmp_path / "report"
    monkeypatch.setenv("NPTC_TX_CHUNK_SIZE", "lots")

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(corrupt_workbook),
            "--report-dir",
            str(report_dir),
            "--check-terminology",
        ],
    )

    assert result.exit_code == 2, result.output
    assert "could not read workbook" in result.output
    assert "invalid terminology configuration" not in result.output
    assert not report_dir.exists()


@pytest.mark.req("FR-54")
def test_a_terminology_failure_writes_no_report_and_exits_three(
    tmp_path: Path, clean_bindings_workbook: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The dangerous alternative is a report written anyway: cell defects
    complete, terminology findings silently absent, indistinguishable from a
    run in which every code validated cleanly."""
    report_dir = tmp_path / "report"
    stub = _ContextStub()
    stub.seed_error(
        Operation.EXPAND,
        TerminologyTransportError("connection refused", operation=Operation.EXPAND),
    )
    _install_stub(monkeypatch, stub)

    result = runner.invoke(
        app,
        [
            "run",
            "--workbook",
            str(clean_bindings_workbook),
            "--report-dir",
            str(report_dir),
            "--check-terminology",
        ],
    )

    assert result.exit_code == 3, result.output
    assert not report_dir.exists()
    assert "Traceback" not in result.output
    assert "terminology validation failed" in result.output
