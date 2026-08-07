"""Tests for the FR-97 designation reconciliation pass (issue #28, P0-6).

Every test drives the real ``check_terminology`` -> ``check_designations``
pipeline against a ``StubTerminologyClient`` (NFR-37): hand-building
``SweepResult``/``ConceptDesignations`` objects would let a test drift from
what the sweep actually produces, which is exactly the risk this pass exists
to avoid on the terminology-client side.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from nptc_shared.terminology.errors import OperationOutcomeIssue, TerminologyStatusError
from nptc_shared.terminology.models import AU_LANGUAGE_TAG, Operation, ValidationResult
from nptc_shared.terminology.stub import StubConcept, StubTerminologyClient
from nptc_shared.terminology.sweep import TerminologySweep
from nptc_transform.bands import Band, FindingCode
from nptc_transform.designation_check import DesignationOutcome, check_designations
from nptc_transform.findings import Finding
from nptc_transform.terminology_check import check_terminology
from nptc_transform.workbook import read_workbook

NBSP = chr(0x00A0)  # non-breaking space
ZWSP = chr(0x200B)  # zero width space - Cf, survives str.strip()

# Real, Verhoeff-valid SCTIDs. 122192001, 391483001 and 413450008 are the
# PRD's own Appendix A.10 sample rows; the rest are fixture-only codes minted
# to pass has_valid_check_digit, not a claim about the live terminology.
ROW22_CODE = "122192001"  # Acanthamoeba culture (procedure) - PRD row 22/A.11
ROW29_CODE = "391483001"  # Microscopy (acid fast bacilli) (procedure) - PRD row 29
ROW45_CODE = "413450008"  # Adenovirus nucleic acid detection (procedure) - PRD row 45
DRIFT_CODE = "200000004"
OTHER_CONCEPT_TARGET_CODE = "300000003"
OTHER_CONCEPT_SOURCE_CODE = "300000019"
NO_MATCH_CODE = "400000002"
INT_ONLY_CODE = "500000009"
SYNONYM_CODE = "500000013"
ABSENT_CODE = "1393151000168101"
MALFORMED_CODE = "12345678"
EDGE_WHITESPACE_CODE = "600000005"
INTERIOR_INVISIBLE_CODE = "600000014"
BLANK_LABEL_CODE = "600000022"
MISSING_LABEL_CODE = "600000033"
DOWNGRADE_CODE = "600000046"

HEADERS = [
    "RCPA Preferred term",
    "Terminology binding (SNOMED CT-AU)",
    "SNOMED CT Fully Specified Name",
]


def _write_text_cell(worksheet: Worksheet, row: int, column: int, value: str) -> None:
    cell = worksheet.cell(row=row, column=column, value=value)
    cell.data_type = "s"


def _not_found(code: str) -> TerminologyStatusError:
    return TerminologyStatusError(
        f"CodeSystem/$lookup returned 404 for {code}",
        operation=Operation.LOOKUP,
        status_code=404,
        issues=(OperationOutcomeIssue(severity="error", code="not-found"),),
    )


def _workbook(tmp_path: Path, rows: list[tuple[str, str | None]]) -> Path:
    """One row per ``(code, label)`` pair; ``label=None`` omits the FSN cell
    entirely, reproducing an empty cell rather than a blank string."""
    path = tmp_path / "designations.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    for index, (code, label) in enumerate(rows, start=2):
        sheet.cell(row=index, column=1, value=f"Term {code}")
        _write_text_cell(sheet, index, 2, code)
        if label is not None:
            sheet.cell(row=index, column=3, value=label)
    workbook.save(path)
    return path


def _outcome(workbook: Path, client: StubTerminologyClient) -> DesignationOutcome:
    sheets = read_workbook(workbook)
    sweep = TerminologySweep(client)
    terminology = check_terminology(sheets, sweep=sweep)
    return check_designations(
        sheets, sweep=sweep, bindings=terminology.bindings, results=terminology.results
    )


def _findings_for(outcome: DesignationOutcome, code: str) -> list[Finding]:
    return [f for f in outcome.findings if f"'{code}'" in f.message]


def _clean(code: str, fsn: str, label: str, **kwargs: object) -> StubConcept:
    """A concept whose AU preferred term equals its own untagged FSN/label -
    so a benign axis-1 outcome doesn't spuriously pick up an axis-2
    preferred-term-differs finding just because no test bothered to seed one.
    Only ``ROW45_CODE`` deliberately omits this, to exercise axis 2."""
    return StubConcept(
        code=code,
        fsn=fsn,
        preferred_terms={AU_LANGUAGE_TAG: label},
        **kwargs,  # type: ignore[arg-type]
    )


@pytest.fixture()
def client() -> StubTerminologyClient:
    return StubTerminologyClient(
        concepts=[
            StubConcept(
                code=ROW22_CODE,
                fsn="Acanthamoeba culture (procedure)",
                synonyms=("Acanthamoeba spp culture",),
            ),
            _clean(
                ROW29_CODE,
                "Microscopy (acid fast bacilli) (procedure)",
                "Microscopy (acid fast bacilli)",
            ),
            StubConcept(
                code=ROW45_CODE,
                fsn="Adenovirus nucleic acid detection (procedure)",
                preferred_terms={AU_LANGUAGE_TAG: "Adenovirus nucleic acid assay"},
            ),
            _clean(DRIFT_CODE, "Drift concept (procedure)", "Drift concept"),
            StubConcept(
                code=SYNONYM_CODE,
                fsn="Synonym concept (procedure)",
                synonyms=("Old synonym label",),
                preferred_terms={AU_LANGUAGE_TAG: "Old synonym label"},
            ),
            _clean(OTHER_CONCEPT_TARGET_CODE, "Target concept (procedure)", "Target concept"),
            StubConcept(code=OTHER_CONCEPT_SOURCE_CODE, fsn="Source concept (procedure)"),
            StubConcept(code=NO_MATCH_CODE, fsn="No match concept (procedure)"),
            StubConcept(
                code=INT_ONLY_CODE, fsn="International only concept (procedure)", editions=("int",)
            ),
            StubConcept(code=ABSENT_CODE, fsn="Absent concept (procedure)", editions=()),
            _clean(
                EDGE_WHITESPACE_CODE,
                "Edge whitespace concept (procedure)",
                "Edge whitespace concept",
            ),
            StubConcept(code=INTERIOR_INVISIBLE_CODE, fsn="Interior invisible concept (procedure)"),
            StubConcept(code=BLANK_LABEL_CODE, fsn="Blank label concept (procedure)"),
            StubConcept(code=MISSING_LABEL_CODE, fsn="Missing label concept (procedure)"),
            _clean(
                DOWNGRADE_CODE,
                "Downgrade concept (procedure)",
                "Downgrade concept",
            ),
        ],
        resolved_version={
            "au": "http://snomed.info/sct/32506021000036107/version/20260531",
            "int": "http://snomed.info/sct/900000000000207008/version/20260501",
        },
    )


# -- axis 1: the four outcomes ------------------------------------------------


@pytest.mark.req("FR-97")
def test_a_label_matching_the_tag_stripped_fsn_produces_no_finding(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(tmp_path, [(DRIFT_CODE, "Drift concept")])
    outcome = _outcome(workbook, client)

    assert outcome.findings == ()
    assert not [r for r in client.requests if r.operation is Operation.CODE_SYSTEM_VALIDATE_CODE]


@pytest.mark.req("FR-97")
def test_a_label_matching_the_fsn_with_its_tag_intact_produces_no_finding(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """Regression: a label equal to the FULL, tagged FSN used to fall through
    to ``LABEL_DESIGNATION_DRIFT`` with a message asserting "is not the FSN
    of this code" - false, since it literally is. Unambiguously the
    concept's own FSN, tag included, so axis 1 must treat it the same as the
    tag-stripped match: no axis-1 finding at all. (Axis 2 is independent and
    correctly still fires here: the label, tag included, is not the same
    string as the current AU preferred term "Drift concept".)"""
    workbook = _workbook(tmp_path, [(DRIFT_CODE, "Drift concept (procedure)")])
    outcome = _outcome(workbook, client)

    axis_one_codes = {
        FindingCode.LABEL_DESIGNATION_DRIFT,
        FindingCode.LABEL_BOUND_TO_OTHER_CONCEPT,
        FindingCode.LABEL_MATCHES_NO_DESIGNATION,
    }
    findings = _findings_for(outcome, DRIFT_CODE)
    assert not [f for f in findings if f.code in axis_one_codes]


@pytest.mark.req("FR-97")
@pytest.mark.req("NFR-38")
def test_a_label_matching_a_valid_synonym_is_informational_and_does_not_block(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """NFR-38 test 14's second half: a label matching a valid synonym on the
    bound concept must not abort the import."""
    workbook = _workbook(tmp_path, [(SYNONYM_CODE, "Old synonym label")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, SYNONYM_CODE)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.LABEL_DESIGNATION_DRIFT
    assert findings[0].band is Band.INFORMATIONAL
    assert "Synonym concept (procedure)" in findings[0].message  # the served FSN


@pytest.mark.req("FR-97")
def test_a_label_matching_a_designation_of_a_different_bound_concept_blocks(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(
        tmp_path,
        [
            (OTHER_CONCEPT_TARGET_CODE, "Target concept"),
            # Transcription error: this row's label is the OTHER code's FSN.
            (OTHER_CONCEPT_SOURCE_CODE, "Target concept"),
        ],
    )
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, OTHER_CONCEPT_SOURCE_CODE)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.LABEL_BOUND_TO_OTHER_CONCEPT
    assert findings[0].band is Band.DATA_DEFECT
    assert OTHER_CONCEPT_TARGET_CODE in findings[0].message
    assert not _findings_for(outcome, OTHER_CONCEPT_TARGET_CODE)


@pytest.mark.req("FR-97")
@pytest.mark.req("NFR-38")
def test_row_22_matching_no_designation_aborts_the_import(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """NFR-38 test 14's first half, the PRD's own row 22 fixture."""
    workbook = _workbook(tmp_path, [(ROW22_CODE, "Acanthamoeba species culture")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, ROW22_CODE)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.LABEL_MATCHES_NO_DESIGNATION
    assert findings[0].band is Band.DATA_DEFECT
    assert "did not match any designation" in findings[0].message


@pytest.mark.req("FR-97")
def test_row_22s_defect_is_not_also_reported_as_a_preferred_term_difference(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """PRD Appendix A.11's arithmetic: 8 drift-only rows plus 1 defect row,
    never 9 rows appearing on both lists."""
    workbook = _workbook(tmp_path, [(ROW22_CODE, "Acanthamoeba species culture")])
    outcome = _outcome(workbook, client)

    codes = [f.code for f in _findings_for(outcome, ROW22_CODE)]
    assert codes == [FindingCode.LABEL_MATCHES_NO_DESIGNATION]


# -- axis 2: preferred-term drift, independent of axis 1 ---------------------


@pytest.mark.req("FR-97")
def test_row_45_matches_the_tag_stripped_fsn_and_still_differs_from_the_preferred_term(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """PRD row 45: the label equals the tag-stripped FSN (outcome 1, no axis-1
    finding) while the AU preferred term has since diverged from *both*."""
    workbook = _workbook(tmp_path, [(ROW45_CODE, "Adenovirus nucleic acid detection")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, ROW45_CODE)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.LABEL_DIFFERS_FROM_PREFERRED_TERM
    assert findings[0].band is Band.INFORMATIONAL
    assert "Adenovirus nucleic acid assay" in findings[0].message


@pytest.mark.req("FR-97")
def test_row_29s_double_parenthesised_fsn_strips_exactly_one_group(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """PRD row 29's implementer caution: the correctly rendered label
    legitimately ends in a parenthesised phrase that is part of the term."""
    workbook = _workbook(tmp_path, [(ROW29_CODE, "Microscopy (acid fast bacilli)")])
    outcome = _outcome(workbook, client)

    assert _findings_for(outcome, ROW29_CODE) == []


# -- request discipline: only the delta is probed -----------------------------


@pytest.mark.req("FR-97")
@pytest.mark.req("FR-52")
def test_only_the_delta_is_probed_with_validate_code(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(
        tmp_path,
        [
            (DRIFT_CODE, "Drift concept"),  # outcome 1, local match
            (ROW29_CODE, "Microscopy (acid fast bacilli)"),  # outcome 1, local match
            (NO_MATCH_CODE, "Something else entirely"),  # local miss -> probed
        ],
    )
    _outcome(workbook, client)

    validate_code_requests = [
        r for r in client.requests if r.operation is Operation.CODE_SYSTEM_VALIDATE_CODE
    ]
    assert len(validate_code_requests) == 1


@pytest.mark.req("FR-97")
def test_a_validate_code_confirmation_downgrades_to_informational(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """The probe can only make an outcome more benign, never less: a local
    miss the server nonetheless confirms becomes informational, not blocking."""
    client.seed_validate_code(
        DOWNGRADE_CODE,
        ValidationResult(code=DOWNGRADE_CODE, result=True, display="A server-only synonym"),
        display="A server-only synonym",
    )
    workbook = _workbook(tmp_path, [(DOWNGRADE_CODE, "A server-only synonym")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, DOWNGRADE_CODE)
    drift_findings = [f for f in findings if f.code == FindingCode.LABEL_DESIGNATION_DRIFT]
    assert len(drift_findings) == 1
    assert drift_findings[0].band is Band.INFORMATIONAL
    assert "confirmed" in drift_findings[0].message
    assert not [
        f
        for f in findings
        if f.code
        in (FindingCode.LABEL_BOUND_TO_OTHER_CONCEPT, FindingCode.LABEL_MATCHES_NO_DESIGNATION)
    ]


@pytest.mark.req("FR-97")
def test_a_terminology_failure_during_the_delta_probe_fails_the_run(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """Unlike an absent code in the status pass, a probe failure is never
    folded into a finding - it propagates and fails the whole run (FR-54)."""
    client.seed_error(
        Operation.CODE_SYSTEM_VALIDATE_CODE, _not_found(NO_MATCH_CODE), key=NO_MATCH_CODE
    )
    workbook = _workbook(tmp_path, [(NO_MATCH_CODE, "Something else entirely")])

    with pytest.raises(TerminologyStatusError):
        _outcome(workbook, client)


# -- edge cases: no finding, and no double-report -----------------------------


@pytest.mark.req("FR-97")
def test_a_label_matching_a_designation_in_only_the_international_edition_is_not_a_defect(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    client.seed_error(Operation.LOOKUP, _not_found(INT_ONLY_CODE), key=INT_ONLY_CODE)
    workbook = _workbook(tmp_path, [(INT_ONLY_CODE, "International only concept")])
    outcome = _outcome(workbook, client)

    assert _findings_for(outcome, INT_ONLY_CODE) == []


@pytest.mark.req("FR-97")
def test_a_code_absent_from_every_edition_produces_no_designation_finding(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    client.seed_error(Operation.LOOKUP, _not_found(ABSENT_CODE), key=ABSENT_CODE)
    workbook = _workbook(tmp_path, [(ABSENT_CODE, "Absent concept")])
    outcome = _outcome(workbook, client)

    assert _findings_for(outcome, ABSENT_CODE) == []
    assert not [
        r
        for r in client.requests
        if r.operation is Operation.CODE_SYSTEM_VALIDATE_CODE and r.detail == ABSENT_CODE
    ]
    assert outcome.run.labels_not_reconciled >= 1


@pytest.mark.req("FR-97")
def test_a_malformed_code_cell_produces_no_designation_finding(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(tmp_path, [(MALFORMED_CODE, "Anything at all")])
    outcome = _outcome(workbook, client)

    assert outcome.findings == ()


@pytest.mark.req("FR-97")
def test_a_label_whose_only_defect_is_surrounding_whitespace_is_not_a_designation_defect(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(tmp_path, [(EDGE_WHITESPACE_CODE, f"{NBSP}Edge whitespace concept{NBSP}")])
    outcome = _outcome(workbook, client)

    assert _findings_for(outcome, EDGE_WHITESPACE_CODE) == []


@pytest.mark.req("FR-97")
def test_a_label_containing_an_interior_invisible_character_is_not_double_reported(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(tmp_path, [(INTERIOR_INVISIBLE_CODE, f"Interior invisible concept{ZWSP}")])
    outcome = _outcome(workbook, client)

    assert _findings_for(outcome, INTERIOR_INVISIBLE_CODE) == []


@pytest.mark.req("FR-97")
def test_an_interior_non_breaking_space_does_not_exempt_a_row_from_reconciliation(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """Regression: an interior non-breaking space used to make
    ``find_invisible_characters`` see an invisible character and skip the
    row entirely - silently converting a blocking `LABEL_BOUND_TO_OTHER_CONCEPT`
    into no finding at all, through the same non-blocking `INVISIBLE_CHARACTER`
    path a plain-space version of the same label would never take (H-07)."""
    workbook = _workbook(
        tmp_path,
        [
            (OTHER_CONCEPT_TARGET_CODE, "Target concept"),
            (OTHER_CONCEPT_SOURCE_CODE, f"Target{NBSP}concept"),
        ],
    )
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, OTHER_CONCEPT_SOURCE_CODE)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.LABEL_BOUND_TO_OTHER_CONCEPT
    assert findings[0].band is Band.DATA_DEFECT
    assert OTHER_CONCEPT_TARGET_CODE in findings[0].message


@pytest.mark.req("FR-97")
def test_a_blank_string_label_produces_no_designation_finding(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(tmp_path, [(BLANK_LABEL_CODE, "   ")])
    outcome = _outcome(workbook, client)

    assert _findings_for(outcome, BLANK_LABEL_CODE) == []


@pytest.mark.req("FR-97")
def test_a_missing_label_cell_produces_no_designation_finding(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(tmp_path, [(MISSING_LABEL_CODE, None)])
    outcome = _outcome(workbook, client)

    assert _findings_for(outcome, MISSING_LABEL_CODE) == []


@pytest.mark.req("FR-97")
def test_a_defect_finding_names_the_international_edition_when_au_did_not_resolve_the_code(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """``_preferred_entry``'s fallback: a code that resolved only in the
    International edition still gets a defect finding quoting *that*
    edition's FSN, not a crash or a silently AU-shaped message."""
    client.seed_error(Operation.LOOKUP, _not_found(INT_ONLY_CODE), key=INT_ONLY_CODE)
    workbook = _workbook(tmp_path, [(INT_ONLY_CODE, "Something else entirely")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, INT_ONLY_CODE)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.LABEL_MATCHES_NO_DESIGNATION
    assert "International only concept (procedure)" in findings[0].message
    assert " in int" in findings[0].message


@pytest.mark.req("FR-97")
def test_a_no_match_finding_truncates_a_long_server_message(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    long_message = "x" * 500
    client.seed_validate_code(
        NO_MATCH_CODE,
        ValidationResult(
            code=NO_MATCH_CODE,
            result=False,
            display="Something else entirely",
            message=long_message,
        ),
        display="Something else entirely",
    )
    workbook = _workbook(tmp_path, [(NO_MATCH_CODE, "Something else entirely")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, NO_MATCH_CODE)
    assert len(findings) == 1
    assert long_message not in findings[0].message
    assert "…" in findings[0].message


@pytest.mark.req("FR-97")
def test_a_no_match_finding_omits_the_server_clause_when_the_server_gives_no_message(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    client.seed_validate_code(
        NO_MATCH_CODE,
        ValidationResult(code=NO_MATCH_CODE, result=False, display="Something else entirely"),
        display="Something else entirely",
    )
    workbook = _workbook(tmp_path, [(NO_MATCH_CODE, "Something else entirely")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, NO_MATCH_CODE)
    assert len(findings) == 1
    assert "the server reported" not in findings[0].message


# -- message hygiene -----------------------------------------------------------


@pytest.mark.req("NFR-38")
def test_a_designation_drift_finding_escapes_invisible_characters_in_the_quoted_label(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(tmp_path, [(SYNONYM_CODE, f"{NBSP}Old synonym label{NBSP}")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, SYNONYM_CODE)
    assert len(findings) == 1
    assert NBSP not in findings[0].message
    assert "<U+00A0>" in findings[0].message


@pytest.mark.req("NFR-38")
def test_a_no_match_finding_escapes_invisible_characters_in_the_server_message(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    client.seed_validate_code(
        NO_MATCH_CODE,
        ValidationResult(
            code=NO_MATCH_CODE,
            result=False,
            display="Something else entirely",
            message=f"did not match{NBSP}anything",
        ),
        display="Something else entirely",
    )
    workbook = _workbook(tmp_path, [(NO_MATCH_CODE, "Something else entirely")])
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, NO_MATCH_CODE)
    assert len(findings) == 1
    assert NBSP not in findings[0].message
    assert "<U+00A0>" in findings[0].message


# -- other-concept ordering and provenance ------------------------------------


@pytest.mark.req("FR-97")
@pytest.mark.req("FR-73")
def test_the_other_concept_list_is_sorted(tmp_path: Path, client: StubTerminologyClient) -> None:
    """A label ambiguous between two *other* bound concepts must name both,
    in a deterministic order (FR-73) - never dict/set iteration order."""
    shared_target_1, shared_target_2, shared_source = "400000018", "400000025", "400000039"
    client.add_concept(
        StubConcept(
            code=shared_target_1, fsn="Target One (procedure)", synonyms=("Ambiguous shared term",)
        )
    )
    client.add_concept(
        StubConcept(
            code=shared_target_2, fsn="Target Two (procedure)", synonyms=("Ambiguous shared term",)
        )
    )
    client.add_concept(StubConcept(code=shared_source, fsn="Source concept (procedure)"))
    workbook = _workbook(
        tmp_path,
        [
            (shared_target_1, "Target One"),
            (shared_target_2, "Target Two"),
            (shared_source, "Ambiguous shared term"),
        ],
    )
    outcome = _outcome(workbook, client)

    findings = _findings_for(outcome, shared_source)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.LABEL_BOUND_TO_OTHER_CONCEPT
    message = findings[0].message
    assert shared_target_1 in message
    assert shared_target_2 in message
    assert message.index(shared_target_1) < message.index(shared_target_2)


@pytest.mark.req("FR-97")
def test_the_report_records_reconciled_and_not_reconciled_and_confirmation_counts(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    workbook = _workbook(
        tmp_path,
        [
            (DRIFT_CODE, "Drift concept"),  # reconciled, local match
            (NO_MATCH_CODE, "Something else entirely"),  # reconciled, probed
            (MISSING_LABEL_CODE, None),  # not reconciled: no label
            (MALFORMED_CODE, "Anything"),  # never checkable at all
        ],
    )
    outcome = _outcome(workbook, client)

    assert outcome.run.labels_reconciled == 2
    assert outcome.run.labels_not_reconciled == 1
    assert outcome.run.label_confirmations == 1
