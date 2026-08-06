"""Shared fixtures for transform/tests.

Also carries the NFR-37 guard: nothing under ``transform/tests`` may open a
real socket. ``ci.yml``'s ``transform-offline`` job proves the same thing with
``iptables``, but only in CI and only after the fact; this fails locally, at
the exact call, with a message naming the requirement - which is what stops a
network-dependent test being written in the first place. Its twin lives in
``shared/tests/conftest.py``.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import httpx
import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet


@pytest.fixture(autouse=True)
def _no_real_network(monkeypatch: pytest.MonkeyPatch) -> None:
    def _refuse(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("transform/tests must not make a real HTTP request (NFR-37)")

    monkeypatch.setattr(httpx.HTTPTransport, "handle_request", _refuse)


# FR-63's documented published header layout.
HEADERS = [
    "RCPA Preferred term",
    "RCPA Synonyms",
    "Usage guidance",
    "Length",
    "Discipline",
    "Subgroup",
    "Specimen",
    "Terminology binding (SNOMED CT-AU)",
    "SNOMED CT Fully Specified Name",
    "Version",
    "History",
]

NBSP = chr(0x00A0)  # non-breaking space
NNBSP = chr(0x202F)  # narrow no-break space
ZWSP = chr(0x200B)  # zero width space - Cf, not Zs, so no deterministic repair


def _write_text_cell(worksheet: Worksheet, row: int, column: int, value: str) -> None:
    """Writes ``value`` as an explicitly text-typed cell (FR-06/FR-63)."""
    cell = worksheet.cell(row=row, column=column, value=value)
    cell.data_type = "s"


@pytest.fixture(scope="session")
def sample_workbook(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """A minimal, clean .xlsx fixture using the real FR-63 header layout.

    No real SPIA data, built in-process; every cell is clean, so a scan
    against it produces no Appendix A findings - existing FR-70/FR-73 tests
    depend on a run against this fixture succeeding without a layout warning.
    """
    workbook_dir = tmp_path_factory.mktemp("workbook")
    path = workbook_dir / "sample.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    row = ["Sample test", "", "", 11, "Chemical", "", "Serum", "", "Sample test", 4, ""]
    sheet.append(row)
    _write_text_cell(sheet, 2, 8, "12345678")
    workbook.save(path)

    return path


@pytest.fixture(scope="session")
def annex_a_workbook(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Reproduces each PRD Appendix A.1-A.3 case, at the PRD's own cell addresses.

    Rows follow the PRD's own numbering where practical:

    - Row 2, 4, 16 (leading indices only), 21, 27, 34, 35: trailing U+00A0 on
      the preferred term (A.1).
    - Row 16: *two consecutive* U+00A0 (A.1).
    - Row 38: U+202F followed by U+00A0 (A.1).
    - H16, H17, H21: trailing U+00A0 on a code cell, itself stored as text
      (A.1, plus the code hygiene it doesn't otherwise violate).
    - Rows 34, 40, 41: 16-digit codes stored as text (clean re: A.2).
    - Rows 20, 24, 42: 18-digit codes stored as text (clean re: A.2).
    - A parallel set of rows with the same codes stored as numbers, to
      exercise CODE_CELL_NOT_TEXT and NUMERIC_PRECISION_RISK together.
    - 45 FSN rows with leading+trailing whitespace, 1 with leading only
      (row 15's pattern), 4 clean (A.3).
    - One fully clean row, to prove a clean cell produces no finding.
    - One row with a trailing U+00A0 in the synonym cell - A.1 names synonym
      cells, alongside preferred terms and code cells, as carrying this
      defect.
    """
    workbook_dir = tmp_path_factory.mktemp("annex_a")
    path = workbook_dir / "annex_a.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(HEADERS)

    def add_row(
        preferred_term: str,
        code: str | int,
        fsn: str,
        *,
        code_as_text: bool = True,
        synonyms: str | None = None,
    ) -> int:
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=preferred_term)
        if synonyms is not None:
            sheet.cell(row=row, column=2, value=synonyms)
        if code_as_text:
            _write_text_cell(sheet, row, 8, str(code))
        else:
            sheet.cell(row=row, column=8, value=code)
        sheet.cell(row=row, column=9, value=fsn)
        return row

    # A.1: trailing U+00A0 on the preferred term.
    add_row(f"Aciclovir level{NBSP}", "111111111", "Aciclovir level")
    # A.1: two consecutive U+00A0.
    add_row(f"Term with double space{NBSP}{NBSP}", "222222222", "Term with double space")
    # A.1: U+202F followed by U+00A0.
    add_row(f"Term with mixed space{NNBSP}{NBSP}", "333333333", "Term with mixed space")
    # A.1: trailing U+00A0 on a text-typed code cell.
    add_row("Clean term one", f"121309009{NBSP}", "Clean term one")

    # A.2: 16-digit and 18-digit codes, stored as text - clean.
    add_row("Clean sixteen digit code", "1393151000168101", "Clean sixteen digit code")
    add_row("Clean eighteen digit code", "933434771000036107", "Clean eighteen digit code")

    # A.2: the same shape, stored as a number - CODE_CELL_NOT_TEXT and
    # NUMERIC_PRECISION_RISK both fire.
    add_row(
        "Sixteen digit code as number",
        1393151000168101,
        "Sixteen digit code as number",
        code_as_text=False,
    )
    add_row(
        "Eighteen digit code as number",
        933434771000036107,
        "Eighteen digit code as number",
        code_as_text=False,
    )
    # A short code stored as a number - CODE_CELL_NOT_TEXT only, no precision risk.
    add_row("Short code as number", 12345678, "Short code as number", code_as_text=False)

    # A.3: leading and trailing whitespace on the FSN.
    add_row("Trichloroethane measurement", "444444444", " 1,1,1-Trichloroethane measurement ")
    # A.3: leading whitespace only.
    add_row(
        "Hydroxymandelate measurement",
        "555555555",
        " 3-Methyl,4-hydroxymandelate measurement",
    )

    # Fully clean row - must produce no finding at all.
    add_row("Clean row", "666666666", "Clean row")

    # A.1: trailing U+00A0 in a synonym cell - the PRD names synonym cells
    # alongside preferred terms and code cells as carrying this defect.
    add_row(
        "Term with synonym space",
        "777777777",
        "Term with synonym space",
        synonyms=f"Synonym one{NBSP}",
    )

    # FR-71 requires-human-decision: a non-Zs invisible character has no
    # single deterministic repair.
    add_row(f"Term with zero width{ZWSP}", "888888888", "Term with zero width")
    # FR-71 requires-human-decision: a synonym cell that's nothing but
    # whitespace - stripping it would empty it entirely.
    add_row(
        "Term with blank synonym",
        "999999999",
        "Term with blank synonym",
        synonyms="   ",
    )

    # FR-71 data-defect: a code cell holding a type with no deterministic
    # coercion to a valid SCTID (unlike a number, there is no value to
    # recover - only a wrong one to report).
    add_row("Term with boolean code", True, "Term with boolean code", code_as_text=False)
    add_row("Term with formula code", "=1+1", "Term with formula code", code_as_text=False)

    workbook.save(path)
    return path


@pytest.fixture(scope="session")
def auto_correctable_only_workbook(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """A workbook whose only defect is auto-correctable (FR-71) - no finding
    in any other band. Distinct from ``sample_workbook``, which is fully
    clean: this fixture exists to prove a run can carry findings and still
    exit 0, because none of them block the import.
    """
    workbook_dir = tmp_path_factory.mktemp("auto_correctable_only")
    path = workbook_dir / "auto_correctable_only.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    sheet.append([f"Aciclovir level{NBSP}", "", "", 11, "Chemical", "", "Serum", "", "x", 4, ""])
    _write_text_cell(sheet, 2, 8, "12345678")
    workbook.save(path)

    return path


@pytest.fixture(scope="session")
def unrecognised_layout_workbook(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """A sheet that resolves *some* SPIA columns but not the code column.

    This is genuine FR-63 layout drift - the code header itself has drifted
    on an otherwise-recognisable SPIA sheet - not a sheet that isn't SPIA
    data at all (that's ``no_spia_columns_workbook`` below). Every row on a
    sheet like this went unscanned for Appendix A.2/A.3 defects, which is why
    it's a blocking data defect rather than an informational notice.
    """
    workbook_dir = tmp_path_factory.mktemp("unrecognised_layout")
    path = workbook_dir / "unrecognised_layout.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Some Column"])
    sheet.append(["value", "value"])
    workbook.save(path)

    return path


@pytest.fixture(scope="session")
def total_header_drift_workbook(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """A named ``Requesting`` data sheet whose header row has drifted so
    completely that it resolves *zero* SPIA column roles - for example, a
    banner row inserted above the real FR-63 headers.

    This produces the identical "no roles resolved" signal
    ``no_spia_columns_workbook`` does, but it is not FR-63's documented
    ``Rev History`` case - it is a genuine SPIA data sheet whose layout has
    drifted, so it must still block the import (``UNRECOGNISED_LAYOUT``,
    data defect), not be waved through as informational just because no
    column happened to be recognised.
    """
    workbook_dir = tmp_path_factory.mktemp("total_header_drift")
    path = workbook_dir / "total_header_drift.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Company confidential", "Do not distribute"])
    sheet.append(["value", "value"])
    workbook.save(path)

    return path


@pytest.fixture(scope="session")
def no_spia_columns_workbook(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """A sheet with no column recognisable as SPIA data at all.

    Reproduces the published workbook's own ``Rev History`` worksheet
    (FR-63, FR-60): hand-written prose, not a drifted SPIA sheet. This is the
    informational, non-blocking outcome of ``_scan_layout`` - the sheet is
    still reported so an operator can see it was skipped, but it must not
    abort the import the way genuine layout drift does.
    """
    workbook_dir = tmp_path_factory.mktemp("no_spia_columns")
    path = workbook_dir / "no_spia_columns.xlsx"

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Rev History"
    sheet.append(["Release notes"])
    sheet.append(["Line one.\nLine two, with a trailing space. "])
    workbook.save(path)

    return path


@pytest.fixture()
def corrupt_workbook(tmp_path: Path) -> Path:
    """A file with an .xlsx extension that is not a valid zip/workbook."""
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not a real workbook")
    return path


@pytest.fixture()
def corrupt_worksheet_workbook(tmp_path: Path) -> Path:
    """A valid zip and workbook manifest, but a sheet XML that's unparsable
    from its very first byte.

    ``ReadOnlyWorksheet.__init__`` parses dimensions eagerly, via a fresh
    ``iterparse`` over the sheet XML (``_get_size`` -> ``parse_dimensions``,
    openpyxl/worksheet/_reader.py) - so if the corruption is at the start of
    the file, ``load_workbook()`` itself fails, before ``iter_rows()`` is
    ever called. Distinct from ``corrupt_row_workbook`` below, where the
    corruption surfaces only once the sheet body is actually iterated.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["A"])
    sheet.append(["x"])
    good_path = tmp_path / "good.xlsx"
    workbook.save(good_path)

    corrupt_path = tmp_path / "corrupt_worksheet.xlsx"
    with (
        zipfile.ZipFile(good_path) as source,
        zipfile.ZipFile(corrupt_path, "w") as dest,
    ):
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = b"<not-valid-xml"
            dest.writestr(item, data)

    return corrupt_path


@pytest.fixture()
def corrupt_row_workbook(tmp_path: Path) -> Path:
    """A valid zip, manifest and sheet dimension - but a malformed later row.

    ``parse_dimensions`` finds the sheet's ``<dimension>`` element (which
    openpyxl always writes before ``<sheetData>``) and returns immediately,
    without reading as far as ``<sheetData>`` at all - so ``load_workbook()``
    succeeds. ``iter_rows()`` re-reads the same XML from the start via a
    second, separate parse and gets as far as yielding the clean earlier
    rows before it reaches the corrupt one. This is the lazy failure path
    ``read_workbook``'s merged except block is also written to catch.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["A", "B"])
    sheet.append(["x", "y"])
    sheet.append(["p", "q"])
    good_path = tmp_path / "good.xlsx"
    workbook.save(good_path)

    with zipfile.ZipFile(good_path) as source:
        sheet_xml = source.read("xl/worksheets/sheet1.xml").decode("utf-8")

    bad_row = (
        '<row r="3"><c r="A3" t="inlineStr"><is><t>p</t></is></c>'
        '<c r="B3" t="inlineStr"><is><t>q</t></is></c></row>'
    )
    assert bad_row in sheet_xml, "openpyxl's row XML shape changed - update bad_row to match"
    corrupted_xml = sheet_xml.replace(bad_row, '<row r="3"><c r="A3"<not-valid')

    corrupt_path = tmp_path / "corrupt_row.xlsx"
    with (
        zipfile.ZipFile(good_path) as source,
        zipfile.ZipFile(corrupt_path, "w") as dest,
    ):
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = corrupted_xml.encode("utf-8")
            dest.writestr(item, data)

    return corrupt_path


@pytest.fixture()
def numeric_overflow_workbook(tmp_path: Path) -> Path:
    """A numeric cell whose raw XML text overflows a double when parsed.

    ``float("1E400")`` returns ``inf`` without raising - Python's own
    ``_cast_number`` behaviour, not a malformed file in any sense
    ``read_workbook`` would reject - so this is a well-formed workbook that
    the reader must read successfully; it's the scanner's ``_digit_count``
    that has to cope with the resulting ``inf`` without crashing.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)"])
    sheet.append([12345678])
    good_path = tmp_path / "good.xlsx"
    workbook.save(good_path)

    with zipfile.ZipFile(good_path) as source:
        sheet_xml = source.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert "<v>12345678</v>" in sheet_xml, "openpyxl's cell XML shape changed"
    overflowed_xml = sheet_xml.replace("<v>12345678</v>", "<v>1E400</v>")

    overflow_path = tmp_path / "numeric_overflow.xlsx"
    with (
        zipfile.ZipFile(good_path) as source,
        zipfile.ZipFile(overflow_path, "w") as dest,
    ):
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = overflowed_xml.encode("utf-8")
            dest.writestr(item, data)

    return overflow_path
