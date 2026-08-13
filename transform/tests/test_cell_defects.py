"""Tests for PRD Appendix A.1-A.3 cell defect detection (FR-70)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from typer.testing import CliRunner

from nptc_transform.bands import Band
from nptc_transform.cell_defects import scan_workbook, split_synonyms
from nptc_transform.cli import app
from nptc_transform.findings import Finding
from nptc_transform.workbook import Sheet, read_workbook

runner = CliRunner()


def _findings_at(sheets: tuple[Sheet, ...], reference: str) -> list[Finding]:
    findings = scan_workbook(sheets)
    return [f for f in findings if str(f.location) == reference]


@pytest.mark.req("FR-70")
def test_clean_workbook_produces_no_findings(sample_workbook: Path) -> None:
    sheets = read_workbook(sample_workbook)
    assert scan_workbook(sheets) == ()


@pytest.mark.req("FR-70")
def test_trailing_nbsp_on_preferred_term_flagged_invisible_character(
    annex_a_workbook: Path,
) -> None:
    sheets = read_workbook(annex_a_workbook)
    findings = _findings_at(sheets, "Requesting!A2")
    codes = {f.code for f in findings}
    assert "INVISIBLE_CHARACTER" in codes
    finding = next(f for f in findings if f.code == "INVISIBLE_CHARACTER")
    assert "U+00A0" in finding.message
    assert chr(0x00A0) not in finding.message


@pytest.mark.req("FR-70")
def test_two_consecutive_nbsp_both_reported_in_offset_order(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    finding = next(
        f for f in _findings_at(sheets, "Requesting!A3") if f.code == "INVISIBLE_CHARACTER"
    )
    # "Term with double space" is 22 characters, so the two U+00A0 sit at 22, 23.
    assert "offset 22" in finding.message
    assert "offset 23" in finding.message
    assert finding.message.index("offset 22") < finding.message.index("offset 23")


@pytest.mark.req("FR-70")
def test_narrow_no_break_space_then_nbsp_both_reported(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    finding = next(
        f for f in _findings_at(sheets, "Requesting!A4") if f.code == "INVISIBLE_CHARACTER"
    )
    assert "U+202F" in finding.message
    assert "U+00A0" in finding.message
    assert finding.message.index("U+202F") < finding.message.index("U+00A0")


@pytest.mark.req("FR-70")
def test_trailing_nbsp_on_code_cell_flagged(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    finding = next(
        f for f in _findings_at(sheets, "Requesting!H5") if f.code == "INVISIBLE_CHARACTER"
    )
    assert "U+00A0" in finding.message


@pytest.mark.req("FR-70")
def test_trailing_nbsp_on_synonym_cell_flagged(annex_a_workbook: Path) -> None:
    """PRD Appendix A.1: non-breaking spaces also appear in synonym cells,
    not just preferred terms and code cells."""
    sheets = read_workbook(annex_a_workbook)
    findings = _findings_at(sheets, "Requesting!B14")
    codes = {f.code for f in findings}
    assert "INVISIBLE_CHARACTER" in codes
    invisible = next(f for f in findings if f.code == "INVISIBLE_CHARACTER")
    assert "U+00A0" in invisible.message
    assert chr(0x00A0) not in invisible.message
    # str.strip() also strips NBSP, so this is a deliberate second finding on
    # the same cell (text.py: different defect classes, different remedies).
    assert "SURROUNDING_WHITESPACE" in codes


@pytest.mark.req("FR-70")
def test_sixteen_and_eighteen_digit_codes_stored_as_text_are_clean(
    annex_a_workbook: Path,
) -> None:
    sheets = read_workbook(annex_a_workbook)
    for reference in ("Requesting!H6", "Requesting!H7"):
        findings = _findings_at(sheets, reference)
        codes = {f.code for f in findings}
        assert "CODE_CELL_NOT_TEXT" not in codes
        assert "NUMERIC_PRECISION_RISK" not in codes


@pytest.mark.req("FR-70")
def test_sixteen_digit_code_as_number_flags_both_defects(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    codes = {f.code for f in _findings_at(sheets, "Requesting!H8")}
    assert "CODE_CELL_NOT_TEXT" in codes
    assert "NUMERIC_PRECISION_RISK" in codes


@pytest.mark.req("FR-70")
def test_eighteen_digit_code_as_number_flags_both_defects(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    codes = {f.code for f in _findings_at(sheets, "Requesting!H9")}
    assert "CODE_CELL_NOT_TEXT" in codes
    assert "NUMERIC_PRECISION_RISK" in codes


@pytest.mark.req("FR-70")
def test_short_code_as_number_flags_type_only_not_precision_risk(
    annex_a_workbook: Path,
) -> None:
    sheets = read_workbook(annex_a_workbook)
    codes = {f.code for f in _findings_at(sheets, "Requesting!H10")}
    assert "CODE_CELL_NOT_TEXT" in codes
    assert "NUMERIC_PRECISION_RISK" not in codes


@pytest.mark.req("FR-71")
@pytest.mark.parametrize("reference", ["Requesting!H17", "Requesting!H18"])
def test_code_cell_holding_a_non_numeric_type_has_no_coercion(
    annex_a_workbook: Path, reference: str
) -> None:
    """A boolean or formula in the code column (H17, H18) is not a number
    that can be deterministically coerced to a string - unlike
    CODE_CELL_NOT_TEXT, there is no correct SCTID to recover, so this must
    land in a blocking band, not auto-correctable."""
    sheets = read_workbook(annex_a_workbook)
    finding = next(f for f in _findings_at(sheets, reference) if f.code == "CODE_CELL_INVALID_TYPE")
    assert finding.band is Band.DATA_DEFECT


@pytest.mark.req("FR-70")
def test_leading_and_trailing_whitespace_on_fsn_flagged(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    finding = next(
        f for f in _findings_at(sheets, "Requesting!I11") if f.code == "SURROUNDING_WHITESPACE"
    )
    assert "leading" in finding.message
    assert "trailing" in finding.message


@pytest.mark.req("FR-70")
def test_leading_only_whitespace_on_fsn_flagged(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    finding = next(
        f for f in _findings_at(sheets, "Requesting!I12") if f.code == "SURROUNDING_WHITESPACE"
    )
    assert "leading" in finding.message
    assert "trailing" not in finding.message


@pytest.mark.req("FR-70")
def test_clean_row_produces_no_finding(annex_a_workbook: Path) -> None:
    sheets = read_workbook(annex_a_workbook)
    assert _findings_at(sheets, "Requesting!A13") == []
    assert _findings_at(sheets, "Requesting!H13") == []
    assert _findings_at(sheets, "Requesting!I13") == []


@pytest.mark.req("FR-70")
def test_unrecognised_layout_flags_the_missing_code_column(
    unrecognised_layout_workbook: Path,
) -> None:
    """Genuine FR-63 layout drift - some SPIA columns resolved, the code
    column didn't - blocks import as a data defect (FR-71) and says how many
    rows on the sheet went unscanned as a result."""
    sheets = read_workbook(unrecognised_layout_workbook)
    findings = scan_workbook(sheets)
    assert any(f.code == "UNRECOGNISED_LAYOUT" for f in findings)
    finding = next(f for f in findings if f.code == "UNRECOGNISED_LAYOUT")
    assert str(finding.location) == "Requesting!A1"
    assert "Some Column" in finding.message
    assert "1 data row(s)" in finding.message
    assert finding.band is Band.DATA_DEFECT


@pytest.mark.req("FR-70")
def test_unrecognised_layout_sheet_gets_no_cell_level_noise(
    unrecognised_layout_workbook: Path,
) -> None:
    """A sheet with a drifted code column isn't cell-scanned - it gets exactly
    the one UNRECOGNISED_LAYOUT finding, not per-cell A.1/A.3 noise on top."""
    sheets = read_workbook(unrecognised_layout_workbook)
    findings = scan_workbook(sheets)
    assert len(findings) == 1
    assert findings[0].code == "UNRECOGNISED_LAYOUT"


@pytest.mark.req("FR-71")
def test_total_header_drift_on_a_named_data_sheet_still_blocks(
    total_header_drift_workbook: Path,
) -> None:
    """A ``Requesting`` sheet whose header row resolves zero SPIA column
    roles (e.g. a banner row inserted above FR-63's real headers) produces
    the same "no roles resolved" signal as a genuinely non-SPIA sheet - but
    it isn't FR-63's documented ``Rev History`` case, so it must still block
    as unrecognised layout rather than being waved through as informational
    just because no column happened to be recognised."""
    sheets = read_workbook(total_header_drift_workbook)
    findings = scan_workbook(sheets)
    assert len(findings) == 1
    assert findings[0].code == "UNRECOGNISED_LAYOUT"
    assert "1 data row(s)" in findings[0].message
    assert findings[0].band is Band.DATA_DEFECT


@pytest.mark.req("FR-71")
def test_no_spia_columns_sheet_is_informational_not_blocking(
    no_spia_columns_workbook: Path,
) -> None:
    """FR-63/FR-60: the published workbook's own ``Rev History`` worksheet is
    a hand-written prose paragraph with no SPIA columns at all. A trailing
    space in its prose and an ALT+ENTER line break inside a paragraph cell
    must not turn into Appendix A findings - only the one layout finding,
    and it must not block the import the way genuine layout drift does."""
    sheets = read_workbook(no_spia_columns_workbook)
    findings = scan_workbook(sheets)
    assert len(findings) == 1
    assert findings[0].code == "SHEET_NOT_SPIA_DATA"
    assert findings[0].band is Band.INFORMATIONAL


@pytest.mark.req("FR-70")
def test_headers_only_sheet_with_no_code_column_is_not_flagged(tmp_path: Path) -> None:
    """No data rows means nothing to validate yet - an empty template sheet
    with an unfamiliar header row is not itself a defect."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Some Column", "Another Column"])
    path = tmp_path / "headers_only.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    findings = scan_workbook(sheets)
    assert findings == ()


@pytest.mark.req("FR-70")
def test_report_never_contains_a_raw_invisible_character(
    tmp_path: Path, annex_a_workbook: Path
) -> None:
    """NFR-38 test 2: no generated output may contain U+00A0 or U+202F, even
    though every finding in this report is about them.

    ``annex_a_workbook`` carries blocking findings (FR-71), so exit 1 is
    expected here - it is not the assertion this test exists to make."""
    report_dir = tmp_path / "report"
    result = runner.invoke(
        app, ["run", "--workbook", str(annex_a_workbook), "--report-dir", str(report_dir)]
    )
    assert result.exit_code == 1, result.output

    for name in ("report.json", "report.md"):
        text = (report_dir / name).read_text(encoding="utf-8")
        assert chr(0x00A0) not in text
        assert chr(0x202F) not in text


@pytest.mark.req("FR-70")
def test_line_break_in_free_text_column_is_not_an_invisible_character(
    tmp_path: Path,
) -> None:
    """An ALT+ENTER line break in a genuine free-text column (Usage guidance)
    on the primary data sheet is ordinary formatting, not an A.1 defect - the
    sheet-scoping fix above only covers non-data sheets; this covers the
    same free-text content sitting *inside* a sheet that has a code column."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)", "Usage guidance"])
    sheet.append(["A term", "12345678", "Line one.\nLine two."])
    path = tmp_path / "guidance.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    findings = scan_workbook(sheets)
    assert findings == ()


@pytest.mark.req("FR-70")
def test_line_break_outside_a_free_text_column_is_still_flagged(tmp_path: Path) -> None:
    """The free-text exemption is scoped to Usage guidance/History - a line
    break inside a preferred term or a code cell is never legitimate and
    must still surface as an INVISIBLE_CHARACTER finding."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.append(["Aciclovir\nlevel", "12345678"])
    path = tmp_path / "preferred_term_break.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    finding = next(
        f
        for f in _findings_at(sheets, "Requesting!A2")
        if f.code == "INVISIBLE_CHARACTER_AMBIGUOUS"
    )
    assert "U+000A" in finding.message
    assert finding.band is Band.REQUIRES_HUMAN_DECISION


@pytest.mark.req("FR-70")
def test_fifteen_digit_number_is_at_the_safe_ceiling_not_flagged(tmp_path: Path) -> None:
    """PRD §2.1: 15 significant decimal digits is Excel's own guaranteed-safe
    ceiling - a 15-digit numeric cell is exactly representable, so it must
    not be reported as a precision risk (only as the wrong storage type)."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)"])
    sheet.append([123456789012345])  # 15 digits
    path = tmp_path / "fifteen.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    codes = {f.code for f in scan_workbook(sheets)}
    assert "NUMERIC_PRECISION_RISK" not in codes
    assert "CODE_CELL_NOT_TEXT" in codes


@pytest.mark.req("FR-70")
def test_sixteen_digit_number_crosses_the_corruption_boundary(tmp_path: Path) -> None:
    """PRD Appendix A.2: 'any SCTID of 16 digits or more... is silently
    corrupted' - the finding must fire exactly at that boundary, not one
    digit later."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)"])
    sheet.append([1393151000168101])  # 16 digits
    path = tmp_path / "sixteen.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    codes = {f.code for f in scan_workbook(sheets)}
    assert "NUMERIC_PRECISION_RISK" in codes


@pytest.mark.req("FR-70")
def test_numeric_precision_risk_fires_outside_the_code_column(tmp_path: Path) -> None:
    """Deliberately column-agnostic (unlike CODE_CELL_NOT_TEXT): a corrupted
    long value sitting in an unexpected column of an otherwise-recognised
    sheet must still surface, not just when it's in the code column."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)", "Version"])
    sheet.append(["12345678", 1393151000168101])  # code column clean; Version corrupted-length
    path = tmp_path / "version_precision.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    codes = {f.code for f in _findings_at(sheets, "Requesting!B2")}
    assert "NUMERIC_PRECISION_RISK" in codes
    assert "CODE_CELL_NOT_TEXT" not in codes


@pytest.mark.req("FR-70")
def test_whitespace_only_cell_gets_a_distinct_message(tmp_path: Path) -> None:
    """A cell that's nothing but whitespace should say so plainly, rather
    than the misleading 'has leading and trailing whitespace' - there's no
    content for either edge to be relative to."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)", "RCPA Synonyms"])
    sheet.append(["12345678", "   "])
    path = tmp_path / "whitespace_only.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    finding = next(
        f for f in _findings_at(sheets, "Requesting!B2") if f.code == "WHITESPACE_ONLY_CELL"
    )
    assert "only whitespace" in finding.message
    assert "leading" not in finding.message
    assert "trailing" not in finding.message
    assert finding.band is Band.REQUIRES_HUMAN_DECISION


@pytest.mark.req("FR-70")
def test_numeric_overflow_is_flagged_not_crashed(numeric_overflow_workbook: Path) -> None:
    """A numeric cell whose raw text overflows a double (openpyxl's own
    ``_cast_number`` returns ``inf`` without raising) must become a finding,
    not an uncaught OverflowError out of ``int(inf)`` - and the message must
    not fabricate a digit count for a value that isn't actually a number."""
    sheets = read_workbook(numeric_overflow_workbook)
    finding = next(f for f in scan_workbook(sheets) if f.code == "NUMERIC_PRECISION_RISK")
    assert "digit" not in finding.message


@pytest.mark.req("FR-70")
def test_a_code_row_with_no_preferred_term_is_flagged_missing(tmp_path: Path) -> None:
    """A row that resolves a code binding but carries no 'RCPA Preferred
    term' value has nothing to seed a designation from - there is no
    deterministic repair, and silently omitting the row (P0-9/#31's own
    reported hazard) is worse than blocking until it is corrected."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.append([None, "12345678"])
    path = tmp_path / "missing_preferred_term.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    finding = next(f for f in scan_workbook(sheets) if f.code == "MISSING_PREFERRED_TERM")
    assert str(finding.location) == "Requesting!B2"
    assert finding.band is Band.DATA_DEFECT


@pytest.mark.req("FR-70")
def test_a_row_with_neither_a_code_nor_a_preferred_term_is_not_flagged(tmp_path: Path) -> None:
    """MISSING_PREFERRED_TERM is about a row that resolves a code binding
    with nothing to seed a designation from - a row with no code binding at
    all is simply not a SPIA data row, and must not be flagged."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.append([None, None])
    path = tmp_path / "blank_row.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    codes = {f.code for f in scan_workbook(sheets)}
    assert "MISSING_PREFERRED_TERM" not in codes
    assert "MISSING_CODE_BINDING" not in codes


@pytest.mark.req("FR-100")
def test_a_preferred_term_row_with_no_code_is_flagged_missing_binding(tmp_path: Path) -> None:
    """A row that carries a 'RCPA Preferred term' value but resolves no code
    binding at all has nothing to bind an entry to (#132) - the mirror of
    MISSING_PREFERRED_TERM for the opposite column."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.append(["Full blood count", None])
    path = tmp_path / "missing_code_binding.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    finding = next(f for f in scan_workbook(sheets) if f.code == "MISSING_CODE_BINDING")
    assert str(finding.location) == "Requesting!A2"
    assert finding.band is Band.DATA_DEFECT


@pytest.mark.req("FR-100")
def test_a_fully_populated_row_is_not_flagged_missing_binding_or_term(tmp_path: Path) -> None:
    """A row with both a preferred term and a code binding must trigger
    neither MISSING_PREFERRED_TERM nor MISSING_CODE_BINDING."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["RCPA Preferred term", "Terminology binding (SNOMED CT-AU)"])
    sheet.append(["Full blood count", "12345678"])
    path = tmp_path / "fully_populated_row.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    codes = {f.code for f in scan_workbook(sheets)}
    assert "MISSING_PREFERRED_TERM" not in codes
    assert "MISSING_CODE_BINDING" not in codes


@pytest.mark.req("FR-89")
def test_any_specimen_does_not_suppress_findings_for_a_co_occurring_named_value(
    tmp_path: Path,
) -> None:
    """cell_defects._scan_specimen must not short-circuit on 'Any': the
    published data is not guaranteed to keep 'Any' from co-occurring with a
    named specimen on one row, and every value in the cell needs a finding
    (or none) reported for it independently, matching what dataset.py seeds."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)", "Specimen"])
    sheet.append(["12345678", "Any; Nasal swab thing"])
    path = tmp_path / "any_and_named.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    codes = [f.code for f in scan_workbook(sheets)]
    assert "SPECIMEN_UNCONSTRAINED_RESOLVED" in codes
    assert "SPECIMEN_VALUE_UNMAPPED" in codes


@pytest.mark.req("FR-04")
def test_a_comma_with_no_following_space_is_not_treated_as_a_synonym_delimiter() -> None:
    """FR-04's whole point is eliminating delimiter hazards - the comma
    fallback must not shatter ordinary SPIA vocabulary like
    '1,25-dihydroxyvitamin D' into two bogus designations just because it
    contains a bare comma with no semicolon in the cell."""
    assert split_synonyms("1,25-dihydroxyvitamin D") == ("1,25-dihydroxyvitamin D",)


@pytest.mark.req("FR-04")
def test_a_comma_space_is_still_a_valid_synonym_delimiter() -> None:
    """PRD Appendix A.10's own documented case: 'ADA RBC, ADA red cells'
    must still split on the comma-space fallback."""
    assert split_synonyms("ADA RBC, ADA red cells") == ("ADA RBC", "ADA red cells")


@pytest.mark.req("FR-88")
def test_specimen_detection_agrees_with_emission_on_an_interior_invisible_character(
    tmp_path: Path,
) -> None:
    """cell_defects._scan_specimen must scan the same corrected text
    dataset.py emits from - otherwise report-only can claim a specimen value
    is unmapped while --emit-dataset resolves it (or the reverse) for the
    identical cell."""
    nbsp = chr(0x00A0)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Requesting"
    sheet.append(["Terminology binding (SNOMED CT-AU)", "Specimen"])
    sheet.append(["12345678", f"whole{nbsp}blood"])
    path = tmp_path / "specimen_nbsp.xlsx"
    workbook.save(path)

    sheets = read_workbook(path)
    codes = {f.code for f in scan_workbook(sheets)}
    assert "SPECIMEN_VALUE_UNMAPPED" not in codes
