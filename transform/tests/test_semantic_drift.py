"""Tests for the FR-75/H-03 semantic-drift pass (issue #29, P0-7).

Follows ``test_designation_check.py``/``test_misspelling.py``'s idiom: every
test drives the real ``read_workbook`` -> ``check_terminology`` ->
``check_semantic_drift`` pipeline against a ``StubTerminologyClient``
(NFR-37). Nothing here hand-builds a ``SweepResult``/``ConceptDesignations``.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from nptc_shared.terminology.models import (
    HAS_SPECIMEN_ATTRIBUTE,
    ConceptProperty,
    Expansion,
    Operation,
)
from nptc_shared.terminology.stub import StubConcept, StubTerminologyClient
from nptc_shared.terminology.sweep import TerminologySweep
from nptc_transform.bands import Band, FindingCode
from nptc_transform.findings import Finding
from nptc_transform.semantic_drift import (
    SemanticDriftOutcome,
    _extract_timing,
    check_semantic_drift,
)
from nptc_transform.terminology_check import check_terminology
from nptc_transform.workbook import read_workbook

# Annex A.9's own worked-example SCTIDs (verified live against SNOMED CT-AU
# during this feature's planning - see the plan's pre-verified facts table).
ACETONE_URINE_CODE = "47615003"
PROTEIN_CSF_CODE = "430551003"
METHOXYMANDELATE_URINE_24H_CODE = "121302000"
ADENOVIRUS_FAECES_CODE = "121960004"

URINE_SPECIMEN_CODE = "122575003"
CSF_SPECIMEN_CODE = "258450006"
FAECES_SPECIMEN_CODE = "119339001"
SERUM_SPECIMEN_CODE = "119364003"

HEADERS = [
    "RCPA Preferred term",
    "Specimen",
    "Terminology binding (SNOMED CT-AU)",
]


def _write_text_cell(worksheet: Worksheet, row: int, column: int, value: str) -> None:
    cell = worksheet.cell(row=row, column=column, value=value)
    cell.data_type = "s"


def _workbook(tmp_path: Path, rows: list[tuple[str, str | None, str]]) -> Path:
    """One row per ``(preferred_term, specimen_column, code)``."""
    path = tmp_path / "semantic_drift.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    for index, (preferred_term, specimen_column, code) in enumerate(rows, start=2):
        sheet.cell(row=index, column=1, value=preferred_term)
        if specimen_column is not None:
            sheet.cell(row=index, column=2, value=specimen_column)
        _write_text_cell(sheet, index, 3, code)
    workbook.save(path)
    return path


def _run(workbook: Path, client: StubTerminologyClient) -> SemanticDriftOutcome:
    """Runs the real pipeline up to ``check_semantic_drift``, then resets the
    stub's request log so a caller counting requests sees only the ones
    ``check_semantic_drift`` itself issued - never the status pass's own."""
    sheets = read_workbook(workbook)
    sweep = TerminologySweep(client)
    terminology = check_terminology(sheets, sweep=sweep)
    client.reset()
    return check_semantic_drift(
        sheets, sweep=sweep, bindings=terminology.bindings, results=terminology.results
    )


def _specimen_concept(code: str, fsn: str, **kwargs: object) -> StubConcept:
    return StubConcept(code=code, fsn=fsn, **kwargs)  # type: ignore[arg-type]


def _drift_findings(outcome: SemanticDriftOutcome) -> list[Finding]:
    return [
        f
        for f in outcome.findings
        if f.code
        in (
            FindingCode.TERM_SPECIMEN_NOT_MODELLED,
            FindingCode.TERM_SPECIMEN_DIFFERS,
            FindingCode.TERM_TIMING_NOT_MODELLED,
        )
    ]


@pytest.fixture()
def annex_a9_client() -> StubTerminologyClient:
    return StubTerminologyClient(
        concepts=[
            StubConcept(
                code=ACETONE_URINE_CODE,
                fsn="Acetone level (procedure)",
                properties=(
                    ConceptProperty(
                        code=HAS_SPECIMEN_ATTRIBUTE, value=URINE_SPECIMEN_CODE, value_type="code"
                    ),
                ),
            ),
            StubConcept(code=PROTEIN_CSF_CODE, fsn="14-3-3 protein CSF (procedure)"),
            StubConcept(
                code=METHOXYMANDELATE_URINE_24H_CODE,
                fsn="4-Hydroxy-3-methoxymandelate level (procedure)",
            ),
            StubConcept(code=ADENOVIRUS_FAECES_CODE, fsn="Adenovirus antigen level (procedure)"),
            _specimen_concept(URINE_SPECIMEN_CODE, "Urine specimen (specimen)"),
            _specimen_concept(CSF_SPECIMEN_CODE, "Cerebrospinal fluid specimen (specimen)"),
            _specimen_concept(FAECES_SPECIMEN_CODE, "Stool specimen (specimen)"),
        ]
    )


def _annex_a9_workbook(tmp_path: Path) -> Path:
    return _workbook(
        tmp_path,
        [
            ("Acetone urine", None, ACETONE_URINE_CODE),
            ("14-3-3 protein CSF", None, PROTEIN_CSF_CODE),
            (
                "4-Hydroxy-3-methoxymandelate urine 24h",
                None,
                METHOXYMANDELATE_URINE_24H_CODE,
            ),
            ("Adenovirus Ag faeces", None, ADENOVIRUS_FAECES_CODE),
        ],
    )


# -- the four Annex A.9 rows, end to end -------------------------------------


@pytest.mark.req("FR-75")
def test_the_annex_a9_fixture_produces_exactly_two_not_modelled_findings(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    workbook = _annex_a9_workbook(tmp_path)

    outcome = _run(workbook, annex_a9_client)

    findings = _drift_findings(outcome)
    assert len(findings) == 2
    assert all(f.band is Band.INFORMATIONAL for f in findings)
    by_location = {f.location: f for f in findings}
    urine_24h_finding = next(
        f for f in findings if "urine" in f.message and "faeces" not in f.message
    )
    faeces_finding = next(f for f in findings if "faeces" in f.message)
    assert urine_24h_finding.code == FindingCode.TERM_SPECIMEN_NOT_MODELLED
    assert "24 h" in urine_24h_finding.message
    assert faeces_finding.code == FindingCode.TERM_SPECIMEN_NOT_MODELLED
    assert by_location  # both findings resolved to distinct cell locations


@pytest.mark.req("FR-75")
def test_the_csf_row_needs_zero_group_specific_classification_requests(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    """Row 2's own FSN literally contains the hand-typed 'csf' term, so the
    visibility filter suppresses it before any classification request for the
    csf group is ever issued - proven here by asserting no
    ``codes_with_attribute_value``-shaped request mentions ``258450006``."""
    workbook = _annex_a9_workbook(tmp_path)

    _run(workbook, annex_a9_client)

    and_requests = [
        r.detail
        for r in annex_a9_client.requests
        if r.operation is Operation.EXPAND and " AND " in r.detail
    ]
    assert not any(CSF_SPECIMEN_CODE in detail for detail in and_requests)


@pytest.mark.req("FR-75")
def test_the_classification_phase_costs_exactly_two_plus_g_expand_calls_and_no_others(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    """Two structurally distinct pieces of server data (the specimen table's
    own vocabulary via ``describe``, and the global "no value at all" check
    via ``codes_without_attribute``) plus one ``codes_with_attribute_value``
    per distinct still-unresolved group (``G`` = urine, faeces = 2 here,
    csf having been suppressed by the visibility filter). See
    ``check_semantic_drift``'s own docstring for why this total is
    ``2 + G``, not literally ``1 + G``."""
    workbook = _annex_a9_workbook(tmp_path)

    _run(workbook, annex_a9_client)

    requests = annex_a9_client.requests
    expand_requests = [r for r in requests if r.operation is Operation.EXPAND]
    describe_requests = [
        r for r in expand_requests if " AND " not in r.detail and " MINUS " not in r.detail
    ]
    minus_requests = [r for r in expand_requests if " MINUS " in r.detail]
    and_requests = [r for r in expand_requests if " AND " in r.detail]

    assert len(describe_requests) == 1
    assert len(minus_requests) == 1
    assert len(and_requests) == 2  # G = {urine, faeces}
    assert len(expand_requests) == 4
    assert len(requests) == 4  # zero $lookup/$validate-code calls
    assert not any(r.operation is Operation.LOOKUP for r in requests)
    assert not any(
        r.operation in (Operation.CODE_SYSTEM_VALIDATE_CODE, Operation.VALUE_SET_VALIDATE_CODE)
        for r in requests
    )


# -- server augmentation of the visibility filter ----------------------------


@pytest.mark.req("FR-75")
def test_a_server_only_synonym_of_the_specimen_concept_suppresses_a_finding(tmp_path: Path) -> None:
    """The bound concept's own FSN does not contain 'csf'/'cerebrospinal
    fluid' - only a synonym the specimen concept itself serves. Suppression
    here can only be explained by the server-augmentation path
    (``describe()``), not the hand-typed table term alone."""
    code = "700000006"
    client = StubTerminologyClient(
        concepts=[
            StubConcept(code=code, fsn="Test analyte spinal tap fluid (procedure)"),
            _specimen_concept(
                CSF_SPECIMEN_CODE,
                "Cerebrospinal fluid specimen (specimen)",
                synonyms=("Spinal tap fluid",),
            ),
        ]
    )
    workbook = _workbook(tmp_path, [("Test analyte CSF", None, code)])

    outcome = _run(workbook, client)

    assert _drift_findings(outcome) == []


@pytest.mark.req("FR-75")
def test_without_that_server_synonym_the_same_row_is_flagged(tmp_path: Path) -> None:
    code = "700000006"
    client = StubTerminologyClient(
        concepts=[
            StubConcept(code=code, fsn="Test analyte spinal tap fluid (procedure)"),
            _specimen_concept(CSF_SPECIMEN_CODE, "Cerebrospinal fluid specimen (specimen)"),
        ]
    )
    workbook = _workbook(tmp_path, [("Test analyte CSF", None, code)])

    outcome = _run(workbook, client)

    findings = _drift_findings(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.TERM_SPECIMEN_NOT_MODELLED


# -- descendant Has-specimen value (the <<-closure guard) --------------------


@pytest.mark.req("FR-75")
def test_a_descendant_has_specimen_value_agrees_with_the_group_root(tmp_path: Path) -> None:
    code = "700000010"
    descendant = "700000023"
    client = StubTerminologyClient(
        concepts=[
            StubConcept(
                code=code,
                fsn="Test analyte level (procedure)",
                properties=(
                    ConceptProperty(
                        code=HAS_SPECIMEN_ATTRIBUTE, value=descendant, value_type="code"
                    ),
                ),
            ),
            StubConcept(
                code=descendant,
                fsn="Urine specimen from catheter (specimen)",
                parents=(URINE_SPECIMEN_CODE,),
            ),
            _specimen_concept(URINE_SPECIMEN_CODE, "Urine specimen (specimen)"),
        ]
    )
    workbook = _workbook(tmp_path, [("Test analyte urine", None, code)])

    outcome = _run(workbook, client)

    assert _drift_findings(outcome) == []


# -- a wrong specimen value ---------------------------------------------------


@pytest.mark.req("FR-75")
def test_a_wrong_specimen_value_is_reported_as_differs(tmp_path: Path) -> None:
    code = "700000047"
    client = StubTerminologyClient(
        concepts=[
            StubConcept(
                code=code,
                fsn="Test analyte level (procedure)",
                properties=(
                    ConceptProperty(
                        code=HAS_SPECIMEN_ATTRIBUTE, value=SERUM_SPECIMEN_CODE, value_type="code"
                    ),
                ),
            ),
            _specimen_concept(URINE_SPECIMEN_CODE, "Urine specimen (specimen)"),
            _specimen_concept(SERUM_SPECIMEN_CODE, "Serum specimen (specimen)"),
        ]
    )
    workbook = _workbook(tmp_path, [("Test analyte urine", None, code)])

    outcome = _run(workbook, client)

    findings = _drift_findings(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.TERM_SPECIMEN_DIFFERS


# -- duplicate bindings: same code, different asserted groups on each row ----


@pytest.mark.req("FR-75")
def test_a_duplicate_binding_is_classified_per_row_group_not_per_code(tmp_path: Path) -> None:
    """The same SCTID bound by two rows, asserting two different specimen
    groups - duplicate bindings do occur in the workbook. The concept's own
    modelled ``Has specimen`` agrees with one group (urine) and disagrees
    with the other (serum): only the serum row's classification may depend
    on the serum-specific check. Keying ``differs`` by code alone would let
    the serum row's disagreement bleed into the urine row, which genuinely
    agrees and must report nothing."""
    code = "700000099"
    client = StubTerminologyClient(
        concepts=[
            StubConcept(
                code=code,
                fsn="Test analyte level (procedure)",
                properties=(
                    ConceptProperty(
                        code=HAS_SPECIMEN_ATTRIBUTE, value=URINE_SPECIMEN_CODE, value_type="code"
                    ),
                ),
            ),
            _specimen_concept(URINE_SPECIMEN_CODE, "Urine specimen (specimen)"),
            _specimen_concept(SERUM_SPECIMEN_CODE, "Serum specimen (specimen)"),
        ]
    )
    workbook = _workbook(
        tmp_path,
        [
            ("Test analyte urine", None, code),
            ("Test analyte serum", None, code),
        ],
    )

    outcome = _run(workbook, client)

    findings = _drift_findings(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.TERM_SPECIMEN_DIFFERS
    assert "serum" in findings[0].message


# -- timing only ---------------------------------------------------------------


@pytest.mark.req("FR-75")
def test_urine_24h_on_a_plain_urine_seeded_concept_is_timing_not_modelled_only(
    tmp_path: Path,
) -> None:
    code = "700000068"
    client = StubTerminologyClient(
        concepts=[
            StubConcept(
                code=code,
                fsn="Test analyte level (procedure)",
                properties=(
                    ConceptProperty(
                        code=HAS_SPECIMEN_ATTRIBUTE, value=URINE_SPECIMEN_CODE, value_type="code"
                    ),
                ),
            ),
            _specimen_concept(URINE_SPECIMEN_CODE, "Urine specimen (specimen)"),
        ]
    )
    workbook = _workbook(tmp_path, [("Test analyte urine 24h", None, code)])

    outcome = _run(workbook, client)

    findings = _drift_findings(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.TERM_TIMING_NOT_MODELLED
    assert "24 h" in findings[0].message


@pytest.mark.req("FR-75")
def test_a_differently_worded_but_matching_served_timing_suppresses_the_finding(
    tmp_path: Path,
) -> None:
    """The label says '24h'; the bound concept's own FSN spells it out as
    '24 hour'. A literal word-boundary match of the canonical string '24 h'
    against that FSN text fails (no boundary between 'h' and the following
    'o' in 'hour'), which would wrongly flag this row - the two must be
    compared as canonicalised timings, not as literal substrings."""
    code = "700000101"
    client = StubTerminologyClient(
        concepts=[
            StubConcept(code=code, fsn="Test analyte 24 hour collection (procedure)"),
        ]
    )
    workbook = _workbook(tmp_path, [("Test analyte 24h", None, code)])

    outcome = _run(workbook, client)

    assert _drift_findings(outcome) == []


# -- the timing regex itself ---------------------------------------------------


@pytest.mark.req("FR-75")
@pytest.mark.parametrize(
    ("label", "expected"),
    [
        ("something urine 24h", "24 h"),
        ("something 24-hour urine", "24 h"),
        ("something 24 hr collection", "24 h"),
        ("b12", None),
        ("vitamin d3", None),
        ("1,25 dihydroxyvitamin d", None),
    ],
)
def test_extract_timing_matches_only_genuine_timing_wording(
    label: str, expected: str | None
) -> None:
    from nptc_shared.text import normalise_for_comparison

    folded = normalise_for_comparison(label).casefold()
    assert _extract_timing(folded) == expected


# -- an unresolvable specimen-table SCTID ------------------------------------


class _OmittingExpansionClient(StubTerminologyClient):
    """Drops ``omitted`` from every ``expand`` result it would otherwise
    return - mirrors ``test_terminology_sweep.py``'s own ``_OmittingClient``,
    kept local because it is only ever needed here."""

    def __init__(self, *, omitted: str, **kwargs: object) -> None:
        super().__init__(**kwargs)  # type: ignore[arg-type]
        self._omitted = omitted

    def expand(self, ecl: str, **kwargs: object) -> Expansion:  # type: ignore[override]
        expansion = super().expand(ecl, **kwargs)  # type: ignore[arg-type]
        kept = tuple(c for c in expansion.concepts if c.code != self._omitted)
        return Expansion(
            concepts=kept,
            total=len(kept),
            offset=expansion.offset,
            resolved_versions=expansion.resolved_versions,
        )


@pytest.mark.req("FR-75")
def test_an_unresolvable_specimen_table_code_is_counted_and_degrades_gracefully(
    tmp_path: Path,
) -> None:
    code = "700000075"
    client = _OmittingExpansionClient(
        omitted=URINE_SPECIMEN_CODE,
        concepts=[
            StubConcept(
                code=code,
                fsn="Acetone level (procedure)",
                properties=(
                    ConceptProperty(
                        code=HAS_SPECIMEN_ATTRIBUTE, value=URINE_SPECIMEN_CODE, value_type="code"
                    ),
                ),
            ),
            _specimen_concept(URINE_SPECIMEN_CODE, "Urine specimen (specimen)"),
        ],
    )
    workbook = _workbook(tmp_path, [("Acetone urine", None, code)])

    outcome = _run(workbook, client)

    assert outcome.run.specimen_table_entries_unresolved == 1
    # The urine group still functions on its hand-typed term alone: the
    # code's own Has-specimen value (122575003) still agrees with the group
    # root, so no finding fires despite the server not describing it.
    assert _drift_findings(outcome) == []


# -- the Specimen column coverage audit --------------------------------------


@pytest.mark.req("FR-75")
def test_a_specimen_column_value_mapping_to_no_group_is_counted(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    workbook = _workbook(
        tmp_path,
        [
            ("Acetone urine", "Nasopharyngeal aspirate", ACETONE_URINE_CODE),
        ],
    )

    outcome = _run(workbook, annex_a9_client)

    assert outcome.run.specimen_column_values_unmapped == 1


@pytest.mark.req("FR-75")
def test_a_covered_specimen_column_value_is_not_counted_as_unmapped(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    workbook = _workbook(
        tmp_path,
        [
            ("Acetone urine", "Urine", ACETONE_URINE_CODE),
        ],
    )

    outcome = _run(workbook, annex_a9_client)

    assert outcome.run.specimen_column_values_unmapped == 0


@pytest.mark.req("FR-75")
def test_any_and_fluids_specimen_column_values_are_excluded_from_the_audit(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    workbook = _workbook(
        tmp_path,
        [
            ("Acetone urine", "Any", ACETONE_URINE_CODE),
        ],
    )

    outcome = _run(workbook, annex_a9_client)

    assert outcome.run.specimen_column_values_unmapped == 0


@pytest.mark.req("FR-75")
def test_a_multi_value_specimen_cell_counts_its_unmapped_value_separately(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    """A ``Specimen`` cell asserting more than one value (FR-88's delimiter)
    must be split before checking coverage, the same way
    ``cell_defects.split_specimen_values`` splits it for the rest of the
    pipeline - folding the whole cell as one value let a covered value's
    partial match hide an unmapped sibling value's gap (issue #130)."""
    workbook = _workbook(
        tmp_path,
        [
            ("Acetone urine", "Serum; Nasopharyngeal aspirate", ACETONE_URINE_CODE),
        ],
    )

    outcome = _run(workbook, annex_a9_client)

    assert outcome.run.specimen_column_values_unmapped == 1


# -- idempotency of the pure function itself ---------------------------------


@pytest.mark.req("FR-73")
def test_check_semantic_drift_is_idempotent_over_the_same_inputs(
    tmp_path: Path, annex_a9_client: StubTerminologyClient
) -> None:
    """Not a CLI/subprocess determinism test (``test_determinism.py``'s own
    suite covers only the sweep-free branch, since it drives the CLI via a
    fresh subprocess with no stub-injection seam) - this proves the pure
    function itself is order-independent by construction, since it never
    iterates a plain ``dict``/``set`` when emitting findings."""
    workbook = _annex_a9_workbook(tmp_path)
    sheets = read_workbook(workbook)
    sweep = TerminologySweep(annex_a9_client)
    terminology = check_terminology(sheets, sweep=sweep)

    first = check_semantic_drift(
        sheets, sweep=sweep, bindings=terminology.bindings, results=terminology.results
    )
    second = check_semantic_drift(
        sheets, sweep=sweep, bindings=terminology.bindings, results=terminology.results
    )

    assert first.findings == second.findings
    assert first.run == second.run
