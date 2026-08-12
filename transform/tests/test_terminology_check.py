"""Tests for the transform's terminology pass (FR-52, FR-74, FR-84, FR-99).

Every test here runs against ``StubTerminologyClient``, which opens no socket
(NFR-37): the transform's own test suite must never need a terminology server,
and CI proves it by blocking egress for this tree.
"""

from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from nptc_shared.terminology.errors import OperationOutcomeIssue, TerminologyStatusError
from nptc_shared.terminology.models import (
    PROCEDURE_ROOT_CODE,
    SNOMED_CT_AU,
    SNOMED_CT_INTERNATIONAL,
    Operation,
)
from nptc_shared.terminology.stub import StubConcept, StubTerminologyClient
from nptc_shared.terminology.sweep import TerminologySweep
from nptc_transform.bands import Band, FindingCode, blocks_import
from nptc_transform.findings import Finding
from nptc_transform.pipeline import Mode, run_transform
from nptc_transform.terminology_check import check_terminology
from nptc_transform.workbook import read_workbook

# Real SNOMED CT identifiers, all Verhoeff-valid, chosen so each row exercises
# one outcome. The concept data behind them is fixture data, not a claim about
# the live terminology - except 243120004 |Regime/therapy|, whose tag is the
# whole point of FR-99 and is quoted from PRD Appendix A.10.
GOOD_CODE = "122192001"
REGIME_CODE = "243120004"
AU_ONLY_CODE = "413450008"
ABSENT_CODE = "1393151000168101"
MALFORMED_CODE = "12345678"
INACTIVE_CODE = "873871000168106"
OUT_OF_SCOPE_CODE = "105590001"

HEADERS = [
    "RCPA Preferred term",
    "Terminology binding (SNOMED CT-AU)",
    "SNOMED CT Fully Specified Name",
]


def _not_found(code: str) -> TerminologyStatusError:
    return TerminologyStatusError(
        f"CodeSystem/$lookup returned 404 for {code}",
        operation=Operation.LOOKUP,
        status_code=404,
        issues=(OperationOutcomeIssue(severity="error", code="not-found"),),
    )


def _write_text_cell(worksheet: Worksheet, row: int, column: int, value: str) -> None:
    cell = worksheet.cell(row=row, column=column, value=value)
    cell.data_type = "s"


@pytest.fixture()
def bindings_workbook(tmp_path: Path) -> Path:
    """One row per terminology outcome, codes stored as text (FR-06)."""
    path = tmp_path / "bindings.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    for index, code in enumerate(
        (
            GOOD_CODE,
            REGIME_CODE,
            AU_ONLY_CODE,
            ABSENT_CODE,
            MALFORMED_CODE,
            INACTIVE_CODE,
            OUT_OF_SCOPE_CODE,
        ),
        start=2,
    ):
        sheet.cell(row=index, column=1, value=f"Term {code}")
        _write_text_cell(sheet, index, 2, code)
        sheet.cell(row=index, column=3, value=f"Label {code}")
    workbook.save(path)
    return path


@pytest.fixture()
def client() -> StubTerminologyClient:
    stub = StubTerminologyClient(
        concepts=[
            StubConcept(
                code=GOOD_CODE,
                fsn="Acanthamoeba culture (procedure)",
                parents=(PROCEDURE_ROOT_CODE,),
            ),
            StubConcept(
                code=REGIME_CODE,
                fsn="Regime/therapy (regime/therapy)",
                parents=(PROCEDURE_ROOT_CODE,),
            ),
            StubConcept(
                code=AU_ONLY_CODE,
                fsn="Adenovirus nucleic acid detection (procedure)",
                parents=(PROCEDURE_ROOT_CODE,),
                editions=("au",),
            ),
            # No edition carries it at all - the transcription-defect case.
            StubConcept(code=ABSENT_CODE, fsn="Absent concept (procedure)", editions=()),
            StubConcept(
                code=INACTIVE_CODE,
                fsn="Fixture duplicate concept (procedure)",
                parents=(PROCEDURE_ROOT_CODE,),
                active=False,
            ),
            # Not under 71388002 at all: FR-84's error.
            StubConcept(code=OUT_OF_SCOPE_CODE, fsn="Fixture substance (substance)"),
        ],
        resolved_version={
            "au": "http://snomed.info/sct/32506021000036107/version/20260531",
            "int": "http://snomed.info/sct/900000000000207008/version/20260501",
        },
    )
    for code in (AU_ONLY_CODE, ABSENT_CODE):
        stub.seed_error(Operation.LOOKUP, _not_found(code), key=code)
    return stub


def _findings(workbook: Path, client: StubTerminologyClient) -> dict[str, list[Finding]]:
    outcome = check_terminology(read_workbook(workbook), sweep=TerminologySweep(client))
    grouped: dict[str, list[Finding]] = {}
    for finding in outcome.findings:
        grouped.setdefault(finding.code, []).append(finding)
    return grouped


@pytest.mark.req("FR-74")
def test_every_code_is_validated_against_both_editions_in_batches(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """FR-74 with FR-52's shape: two editions, and the request count scales
    with chunks and the delta - never with the number of code cells."""
    check_terminology(read_workbook(bindings_workbook), sweep=TerminologySweep(client))

    expansions = [request for request in client.requests if request.operation is Operation.EXPAND]
    status = [request for request in expansions if " MINUS " not in request.detail]
    hierarchy = [request for request in expansions if " MINUS " in request.detail]
    assert len(status) == 2  # one chunk, one per edition
    assert len(hierarchy) == 2  # FR-84, exactly one per edition
    assert not [
        request
        for request in client.requests
        if request.operation
        in {Operation.CODE_SYSTEM_VALIDATE_CODE, Operation.VALUE_SET_VALIDATE_CODE}
    ]


@pytest.mark.req("FR-06")
def test_a_code_that_is_not_a_well_formed_sctid_is_reported_and_never_submitted(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """It fails the Verhoeff check digit, so it cannot be a real SCTID -
    FR-71 bands that a data defect. Sending it anyway would report it as
    "not found", which reads as a terminology outcome rather than as the
    transcription defect it is."""
    grouped = _findings(bindings_workbook, client)

    malformed = grouped[FindingCode.CODE_NOT_WELL_FORMED]
    assert [str(finding.location) for finding in malformed] == ["Requesting!B6"]
    assert malformed[0].band is Band.DATA_DEFECT
    assert not [request for request in client.requests if MALFORMED_CODE in request.detail]


@pytest.mark.req("FR-74")
def test_a_code_absent_from_every_edition_is_a_blocking_data_defect(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    grouped = _findings(bindings_workbook, client)

    absent = grouped[FindingCode.CODE_NOT_FOUND]
    assert [str(finding.location) for finding in absent] == ["Requesting!B5"]
    assert absent[0].band is Band.DATA_DEFECT
    assert "au, int" in absent[0].message


@pytest.mark.req("FR-74")
def test_a_code_present_only_in_the_au_edition_is_not_a_finding(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """FR-47 case 1: "expected and correct for Australian extension content".
    The catalogue is bound to SNOMED CT-AU, so reporting every AU extension
    code as missing from International would bury the real findings."""
    grouped = _findings(bindings_workbook, client)

    assert all(
        str(finding.location) != "Requesting!B4"
        for findings in grouped.values()
        for finding in findings
    )


@pytest.mark.req("FR-74")
def test_an_inactive_code_is_a_blocking_data_defect(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    grouped = _findings(bindings_workbook, client)

    inactive = grouped[FindingCode.CODE_INACTIVE]
    assert [str(finding.location) for finding in inactive] == ["Requesting!B7"]
    assert inactive[0].band is Band.DATA_DEFECT


@pytest.mark.req("FR-84")
@pytest.mark.req("NFR-38")
def test_a_code_outside_the_procedure_hierarchy_blocks_publication(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """NFR-38 test 13: the finding is an error that blocks, and the check
    that produced it is one batch request per edition (asserted above)."""
    grouped = _findings(bindings_workbook, client)

    violations = grouped[FindingCode.OUT_OF_SCOPE_HIERARCHY]
    assert [str(finding.location) for finding in violations] == ["Requesting!B8"]
    assert violations[0].band is Band.DATA_DEFECT
    assert "71388002" in violations[0].message


@pytest.mark.req("FR-99")
def test_an_unexpected_semantic_tag_is_a_warning_that_does_not_block(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """The failure mode FR-99 exists to prevent: 243120004 |Regime/therapy|
    *is* subsumed by |Procedure|, so treating its tag as an error would abort
    a seeding run over a valid binding."""
    grouped = _findings(bindings_workbook, client)

    tagged = grouped[FindingCode.UNEXPECTED_SEMANTIC_TAG]
    assert [str(finding.location) for finding in tagged] == ["Requesting!B3"]
    assert tagged[0].band is Band.INFORMATIONAL
    assert not blocks_import(tagged[0].band)
    assert "regime/therapy" in tagged[0].message


@pytest.mark.req("FR-99")
def test_the_tag_warning_is_raised_once_even_though_both_editions_serve_it(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    grouped = _findings(bindings_workbook, client)

    assert len(grouped[FindingCode.UNEXPECTED_SEMANTIC_TAG]) == 1


@pytest.mark.req("FR-99")
def test_an_unexpected_semantic_tag_escapes_invisible_characters_in_the_fsn(
    tmp_path: Path,
) -> None:
    """The served FSN is exactly the kind of text this transform exists to
    surface defects in - an invisible character inside it must be named by
    codepoint, never written raw into the report, the same rule every other
    finding in this module already follows (see the sibling
    CODE_NOT_WELL_FORMED message). Without this, the one character that
    would explain the discrepancy to a reviewer is the one made invisible.
    """
    nbsp = "\u00a0"
    code = REGIME_CODE
    path = tmp_path / "tagged.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    _write_text_cell(sheet, 2, 2, code)
    workbook.save(path)

    stub = StubTerminologyClient(
        concepts=[
            StubConcept(
                code=code,
                fsn=f"Regime{nbsp}therapy (regime/therapy)",
                parents=(PROCEDURE_ROOT_CODE,),
            )
        ]
    )

    outcome = check_terminology(read_workbook(path), sweep=TerminologySweep(stub))

    tagged = [f for f in outcome.findings if f.code == FindingCode.UNEXPECTED_SEMANTIC_TAG]
    assert len(tagged) == 1
    assert nbsp not in tagged[0].message
    assert "<U+00A0>" in tagged[0].message


@pytest.mark.req("FR-74")
def test_a_clean_binding_produces_no_finding_at_all(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    grouped = _findings(bindings_workbook, client)

    assert all(
        str(finding.location) != "Requesting!B2"
        for findings in grouped.values()
        for finding in findings
    )


@pytest.mark.req("FR-06")
def test_a_non_text_code_cell_is_excluded_from_terminology_checking_entirely(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """``cell_defects.py`` already owns the type defect for a non-text code
    cell (``CODE_CELL_INVALID_TYPE`` for a date, ``CODE_CELL_NOT_TEXT`` for a
    number) - this pass must not also submit the cell's rendered text (an
    ISO date string, a plain digit string for a number) as though it were a
    transcribed code, which would either report a second, misleading
    ``CODE_NOT_WELL_FORMED`` finding or - worse - silently validate a code
    the docstring's own de-duplication discipline says belongs solely to the
    cell-type scanner."""
    path = tmp_path / "wrong_types.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    sheet.cell(row=2, column=2, value=datetime.date(2024, 1, 5))  # CellType.DATE
    sheet.cell(row=3, column=2, value=int(GOOD_CODE))  # CellType.NUMBER
    workbook.save(path)

    outcome = check_terminology(read_workbook(path), sweep=TerminologySweep(client))

    assert outcome.findings == ()
    assert outcome.run.codes_checked == 0
    assert outcome.run.codes_not_checked == 2
    assert client.requests == ()


@pytest.mark.req("FR-72")
def test_a_code_bound_by_several_rows_is_reported_against_each_cell(
    tmp_path: Path, client: StubTerminologyClient
) -> None:
    """RCPA-QAP works in cells, not codes: one finding per cell, even though
    the sweep resolved the code once."""
    path = tmp_path / "repeated.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    for row in (2, 3):
        _write_text_cell(sheet, row, 2, OUT_OF_SCOPE_CODE)
    workbook.save(path)

    outcome = check_terminology(read_workbook(path), sweep=TerminologySweep(client))

    assert [str(finding.location) for finding in outcome.findings] == [
        "Requesting!B2",
        "Requesting!B3",
    ]
    assert outcome.run.codes_checked == 1


@pytest.mark.req("FR-48")
def test_the_run_records_the_edition_version_each_sweep_resolved_against(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    outcome = check_terminology(read_workbook(bindings_workbook), sweep=TerminologySweep(client))

    assert [edition.label for edition in outcome.run.editions] == ["au", "int"]
    assert outcome.run.editions[0].resolved_versions == (
        "http://snomed.info/sct/32506021000036107/version/20260531",
    )
    assert outcome.run.codes_checked == 6
    assert outcome.run.codes_not_checked == 1


@pytest.mark.req("FR-74")
def test_a_single_edition_run_is_supported_for_a_targeted_check(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    outcome = check_terminology(
        read_workbook(bindings_workbook), sweep=TerminologySweep(client), editions=[SNOMED_CT_AU]
    )

    assert [edition.label for edition in outcome.run.editions] == ["au"]
    # AU-only content resolves here, so the absent-from-every-edition finding
    # must not fire for it just because International was not consulted.
    assert all(
        str(finding.location) != "Requesting!B4" or finding.code != FindingCode.CODE_NOT_FOUND
        for finding in outcome.findings
    )


@pytest.mark.req("FR-74")
def test_asking_for_no_editions_at_all_is_refused(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """With no edition to be absent from, "absent from every edition" is
    vacuously true - an empty list would report the whole workbook as codes
    that do not exist."""
    with pytest.raises(ValueError, match="at least one edition"):
        check_terminology(
            read_workbook(bindings_workbook), sweep=TerminologySweep(client), editions=()
        )


@pytest.mark.req("FR-74")
def test_duplicate_edition_labels_are_refused(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """A pinned and an unpinned edition sharing a label (FR-49's
    reproduce-a-historical-run case) would otherwise silently collapse in
    the per-edition results dict below - both sweeps run against the shared
    server, but one result is simply discarded, understating what was
    actually validated."""
    with pytest.raises(ValueError, match="distinct edition labels"):
        check_terminology(
            read_workbook(bindings_workbook),
            sweep=TerminologySweep(client),
            editions=[SNOMED_CT_AU, SNOMED_CT_AU.pinned_to("20250630")],
        )


@pytest.mark.req("FR-73")
def test_the_pipeline_runs_without_a_sweep_and_records_that_none_ran(
    bindings_workbook: Path,
) -> None:
    """The default is off: a plain run must neither need the network nor
    imply that the codes were validated."""
    result = run_transform(bindings_workbook, mode=Mode.REPORT_ONLY)

    assert result.terminology is None
    assert not [
        finding
        for finding in result.findings
        if finding.code
        in {
            FindingCode.CODE_NOT_FOUND,
            FindingCode.CODE_INACTIVE,
            FindingCode.OUT_OF_SCOPE_HIERARCHY,
            FindingCode.UNEXPECTED_SEMANTIC_TAG,
        }
    ]


@pytest.mark.req("FR-71")
def test_the_pipeline_merges_terminology_findings_and_blocks_on_them(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    result = run_transform(bindings_workbook, mode=Mode.REPORT_ONLY, sweep=TerminologySweep(client))

    codes = {finding.code for finding in result.findings}
    assert FindingCode.OUT_OF_SCOPE_HIERARCHY in codes
    assert result.has_blocking_findings
    assert result.terminology is not None
    assert result.band_counts[Band.INFORMATIONAL] == 1


@pytest.mark.req("FR-73")
def test_two_runs_over_the_same_workbook_and_server_produce_identical_findings(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    first = run_transform(bindings_workbook, mode=Mode.REPORT_ONLY, sweep=TerminologySweep(client))
    second = run_transform(bindings_workbook, mode=Mode.REPORT_ONLY, sweep=TerminologySweep(client))

    assert first.findings == second.findings
    assert first.terminology == second.terminology


@pytest.mark.req("NFR-37")
def test_the_whole_pass_runs_against_the_stub_with_no_network_access(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """The issue's own acceptance criterion, with conftest's autouse guard
    active: any real HTTP request from this tree fails the test at the call."""
    outcome = check_terminology(read_workbook(bindings_workbook), sweep=TerminologySweep(client))

    assert outcome.findings
    assert client.requests


@pytest.mark.req("FR-74")
def test_editions_of_the_international_edition_alone_report_au_only_content_absent(
    bindings_workbook: Path, client: StubTerminologyClient
) -> None:
    """The mirror of the AU-only test: the same code *is* a finding when
    International is the only edition asked, which is what proves the earlier
    silence comes from combining editions rather than from never checking."""
    outcome = check_terminology(
        read_workbook(bindings_workbook),
        sweep=TerminologySweep(client),
        editions=[SNOMED_CT_INTERNATIONAL],
    )

    assert [
        str(finding.location)
        for finding in outcome.findings
        if finding.code == FindingCode.CODE_NOT_FOUND
    ] == ["Requesting!B4", "Requesting!B5"]
