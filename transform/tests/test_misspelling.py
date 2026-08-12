"""Tests for the FR-79/H-04 misspelling heuristics pass (issue #29, P0-7).

Follows ``test_designation_check.py``'s idiom: every test drives the real
``read_workbook`` -> (optionally) ``check_terminology`` over a
``StubTerminologyClient`` -> ``check_misspellings`` pipeline. Nothing here
hand-builds a ``SweepResult``/``ConceptDesignations`` - that would let a test
drift from what the sweep actually produces. A local ``_workbook`` builder
(mirroring ``test_designation_check.py``'s own) is used instead of a fixed
session fixture, because each scenario below needs genuinely different
row content - a single shared fixture would either not cover them or would
force artificial content into ``conftest.py``'s ``annex_a_workbook``, which
this suite must not touch (its own band-count assertions depend on its exact
findings).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.worksheet import Worksheet

from nptc_shared.similarity import tokenise
from nptc_shared.terminology.stub import StubConcept, StubTerminologyClient
from nptc_shared.terminology.sweep import TerminologySweep
from nptc_transform.bands import Band, FindingCode
from nptc_transform.findings import Finding
from nptc_transform.misspelling import AuthoritySource, MisspellingOutcome, check_misspellings
from nptc_transform.terminology_check import check_terminology
from nptc_transform.workbook import read_workbook

# Real, Verhoeff-valid SCTIDs, minted for this fixture only (not a claim
# about the live terminology) - see test_designation_check.py's own comment
# on the same practice.
EPINEPHRINE_CODE = "700000006"
AMYLASE_CODE = "700000010"
AMYLOSE_CODE = "700000023"
UNRELATED_SERTONIN_CODE = "700000047"

HEADERS = [
    "RCPA Preferred term",
    "RCPA Synonyms",
    "Usage guidance",
    "Length",
    "Discipline",
    "Subgroup",
    "Specimen",
    "Terminology binding (SNOMED CT-AU)",
    "SNOMED CT Fully Specified Name",
    "Version",
    "History",
]


def _write_text_cell(worksheet: Worksheet, row: int, column: int, value: str) -> None:
    cell = worksheet.cell(row=row, column=column, value=value)
    cell.data_type = "s"


def _workbook(tmp_path: Path, rows: list[tuple[str | None, str | None, str | None]]) -> Path:
    """One row per ``(preferred_term, synonyms, code)`` - any may be ``None``
    to omit that cell entirely."""
    path = tmp_path / "misspelling.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Requesting"
    sheet.append(HEADERS)
    for index, (preferred_term, synonyms, code) in enumerate(rows, start=2):
        if preferred_term is not None:
            sheet.cell(row=index, column=1, value=preferred_term)
        if synonyms is not None:
            sheet.cell(row=index, column=2, value=synonyms)
        if code is not None:
            _write_text_cell(sheet, index, 8, code)
    workbook.save(path)
    return path


def _outcome(workbook: Path, client: StubTerminologyClient | None = None) -> MisspellingOutcome:
    sheets = read_workbook(workbook)
    if client is None:
        return check_misspellings(sheets)
    sweep = TerminologySweep(client)
    terminology = check_terminology(sheets, sweep=sweep)
    return check_misspellings(sheets, results=terminology.results)


def _misspelling_codes(outcome: MisspellingOutcome) -> list[Finding]:
    return [
        f
        for f in outcome.findings
        if f.code in (FindingCode.PROBABLE_MISSPELLING, FindingCode.INCONSISTENT_SPELLING)
    ]


# -- must flag -----------------------------------------------------------------


@pytest.mark.req("FR-79")
def test_a_single_row_typo_is_flagged_via_the_served_designation_alone(tmp_path: Path) -> None:
    """Heuristic 1's arm (b): with zero cross-row reasoning at all - only one
    row in the whole workbook - a typo is still caught because it is
    compared against the *served* designation of the concept its own code
    binds to."""
    client = StubTerminologyClient(
        concepts=[StubConcept(code=EPINEPHRINE_CODE, fsn="Epinephrine (substance)")]
    )
    workbook = _workbook(tmp_path, [("Epinephine", None, EPINEPHRINE_CODE)])
    outcome = _outcome(workbook, client)

    findings = _misspelling_codes(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.PROBABLE_MISSPELLING
    assert findings[0].band is Band.INFORMATIONAL
    assert "Epinephine" in findings[0].message
    assert "Epinephrine" in findings[0].message


@pytest.mark.req("FR-79")
def test_a_synonym_loses_to_the_preferred_term_with_no_sweep_at_all(tmp_path: Path) -> None:
    """PRD row 51 verbatim: 'antental' in synonyms loses to 'antenatal' in
    the preferred term - and this works on the workbook alone, with
    ``results=None``, proving the tie-break needs no terminology sweep."""
    workbook = _workbook(tmp_path, [("Antenatal screen", "antental", None)])
    outcome = _outcome(workbook)

    findings = _misspelling_codes(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.PROBABLE_MISSPELLING
    assert str(findings[0].location).endswith("!B2")  # the synonyms cell, not the preferred term
    assert "antental" in findings[0].message
    assert "Antenatal" in findings[0].message


@pytest.mark.req("FR-79")
def test_a_rare_token_flags_against_a_far_more_common_one_with_no_sweep(tmp_path: Path) -> None:
    """Heuristic 2 only: 'Bilirubon' never co-occurs with 'Bilirubin' in any
    one entry, so heuristic 1 has nothing to say about it - only the
    corpus-wide frequency gap (1 entry vs. 3) makes it a probable
    inconsistent spelling."""
    workbook = _workbook(
        tmp_path,
        [
            ("Bilirubon", None, None),
            ("Bilirubin", None, None),
            ("Bilirubin panel", None, None),
            ("Bilirubin ratio", None, None),
        ],
    )
    outcome = _outcome(workbook)

    findings = _misspelling_codes(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.INCONSISTENT_SPELLING
    assert "Bilirubon" in findings[0].message
    assert "Bilirubin" in findings[0].message


@pytest.mark.req("FR-79")
def test_heuristic_one_wins_over_heuristic_two_for_the_same_cell_and_token(
    tmp_path: Path,
) -> None:
    """'Troponon' both has an in-entry reference ('Troponin' in the same
    row's preferred term - heuristic 1) and is rare against a corpus-common
    'Troponin' (heuristic 2). At most one finding per cell/token, and
    heuristic 1 takes precedence: the code must be PROBABLE_MISSPELLING,
    never INCONSISTENT_SPELLING."""
    workbook = _workbook(
        tmp_path,
        [
            ("Troponin", "Troponon", None),
            ("Troponin panel", None, None),
            ("Troponin ratio", None, None),
        ],
    )
    outcome = _outcome(workbook)

    findings = _misspelling_codes(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.PROBABLE_MISSPELLING
    assert "Troponon" in findings[0].message


# -- must not flag --------------------------------------------------------------


@pytest.mark.req("FR-79")
def test_a_short_abbreviation_with_a_digit_is_never_compared_at_all(tmp_path: Path) -> None:
    """'ADA2' carries a digit, so it never passes ``is_comparable_token`` -
    appearing across unrelated entries must not produce a finding."""
    workbook = _workbook(
        tmp_path,
        [
            ("ADA2 assay", None, None),
            ("ADA2 confirmatory assay", None, None),
            ("ADA2 follow-up assay", None, None),
        ],
    )
    outcome = _outcome(workbook)

    assert _misspelling_codes(outcome) == []


@pytest.mark.req("FR-79")
def test_entries_differing_only_by_specimen_are_not_a_heuristic_one_signal(
    tmp_path: Path,
) -> None:
    """Annex A.5's negative controls are near-identical adjacent entries
    differing only by specimen - row adjacency is a deliberate non-feature.
    'Potasium' (typo) has no in-entry reference of its own and must not be
    rescued by the clean 'Potassium' spelling one row above it."""
    workbook = _workbook(
        tmp_path,
        [
            ("Potassium", None, None),
            ("Potasium", None, None),
        ],
    )
    outcome = _outcome(workbook)

    assert _misspelling_codes(outcome) == []


@pytest.mark.req("FR-79")
def test_a_hyphenated_abbreviation_pair_does_not_near_match(tmp_path: Path) -> None:
    """'17-OHP' and '17-OH progesterone' tokenise to short/digit-bearing
    pieces that never reach the comparable-token gate, so this well-known
    pathology abbreviation pair must not collide."""
    assert tokenise("17-OHP") == ("17", "OHP")
    assert tokenise("17-OH progesterone") == ("17", "OH", "progesterone")

    workbook = _workbook(tmp_path, [("17-OHP", None, None), ("17-OH progesterone", None, None)])
    outcome = _outcome(workbook)

    assert _misspelling_codes(outcome) == []


@pytest.mark.req("FR-79")
def test_two_digit_bearing_tokens_in_one_cell_do_not_collide(tmp_path: Path) -> None:
    """'7DHC' and '8DHC' co-occurring in the same synonym cell both carry a
    digit, so the digit gate excludes them from comparison entirely."""
    workbook = _workbook(tmp_path, [("Vitamin D precursor", "7DHC 8DHC", None)])
    outcome = _outcome(workbook)

    assert _misspelling_codes(outcome) == []


@pytest.mark.req("FR-79")
def test_adrenal_ab_produces_no_misspelling_finding(tmp_path: Path) -> None:
    """'Adrenal Ab' colliding with something else is FR-05's collision
    detection concern, not FR-79's - and 'Ab' is too short to ever be a
    comparable token in the first place."""
    workbook = _workbook(tmp_path, [("Adrenal Ab", None, None)])
    outcome = _outcome(workbook)

    assert _misspelling_codes(outcome) == []


@pytest.mark.req("FR-71")
@pytest.mark.req("FR-79")
def test_the_synonym_delimiter_does_not_change_the_finding_count(tmp_path: Path) -> None:
    """FR-71's own unresolved comma-vs-semicolon question (PRD Appendix A.4):
    the same synonym content expressed under each delimiter, and under bare
    spaces, must produce the identical (here: zero) count of findings."""
    comma_dir, semicolon_dir, space_dir = (
        tmp_path / "comma",
        tmp_path / "semicolon",
        tmp_path / "space",
    )
    for directory in (comma_dir, semicolon_dir, space_dir):
        directory.mkdir()
    comma = _workbook(comma_dir, [("Widal test", "ADA RBC, ADA red cells", None)])
    semicolon = _workbook(semicolon_dir, [("Widal test", "ADA RBC; ADA red cells", None)])
    bare_space = _workbook(space_dir, [("Widal test", "ADA RBC ADA red cells", None)])

    counts = {len(_misspelling_codes(_outcome(wb))) for wb in (comma, semicolon, bare_space)}
    assert counts == {0}


# -- principal failure mode: the whitelist's precision trade-off --------------


@pytest.mark.req("FR-79")
def test_two_genuinely_distinct_served_analytes_are_not_flagged_when_both_are_authoritative(
    tmp_path: Path,
) -> None:
    """'Amylase' (enzyme) and 'Amylose' (starch) are both real, both served,
    one edit apart. 'Amylose' is corpus-rare (1 entry) against a
    corpus-common 'Amylase' (3 entries) - exactly heuristic 2's shape - but
    because a completed sweep confirms 'Amylose' is itself a genuine served
    designation, the whitelist must suppress the false flag."""
    client = StubTerminologyClient(
        concepts=[
            StubConcept(code=AMYLASE_CODE, fsn="Amylase (substance)"),
            StubConcept(code=AMYLOSE_CODE, fsn="Amylose (substance)"),
        ]
    )
    workbook = _workbook(
        tmp_path,
        [
            ("Amylose", None, AMYLOSE_CODE),
            ("Amylase", None, AMYLASE_CODE),
            ("Amylase panel", None, None),
            ("Amylase ratio", None, None),
        ],
    )
    outcome = _outcome(workbook, client)

    assert _misspelling_codes(outcome) == []
    assert outcome.run.authority_source is AuthoritySource.SWEEP


@pytest.mark.req("FR-79")
def test_the_same_pair_is_flagged_without_a_sweep_documenting_the_precision_loss(
    tmp_path: Path,
) -> None:
    """Companion to the test above: the identical workbook, with no
    terminology sweep at all, DOES flag 'Amylose' - the honest precision
    loss ``AuthoritySource.WORKBOOK_ONLY`` exists to document, not to hide."""
    workbook = _workbook(
        tmp_path,
        [
            ("Amylose", None, None),
            ("Amylase", None, None),
            ("Amylase panel", None, None),
            ("Amylase ratio", None, None),
        ],
    )
    outcome = _outcome(workbook)

    findings = _misspelling_codes(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.INCONSISTENT_SPELLING
    assert "Amylose" in findings[0].message
    assert outcome.run.authority_source is AuthoritySource.WORKBOOK_ONLY


@pytest.mark.req("FR-79")
def test_a_genuine_typo_that_coincides_with_an_unrelated_served_word_is_not_flagged(
    tmp_path: Path,
) -> None:
    """The inverse, accepted miss. 'Sertonin' here really is a typo of
    'Serotonin' - but it happens to be, coincidentally, exactly the served
    designation of some entirely unrelated concept elsewhere in the
    terminology. The whitelist cannot distinguish "genuinely this word" from
    "coincidentally spelled the same as this word", so it suppresses the
    flag either way. This is accepted and must NOT be "fixed" by narrowing
    the whitelist - doing so would just re-introduce the false positive
    the test above exists to suppress."""
    client = StubTerminologyClient(
        concepts=[StubConcept(code=UNRELATED_SERTONIN_CODE, fsn="Sertonin (substance)")]
    )
    workbook = _workbook(
        tmp_path,
        [
            ("Sertonin", None, None),  # the genuine typo
            ("Serotonin", None, None),
            ("Serotonin panel", None, None),
            ("Serotonin ratio", None, None),
            ("Unrelated filler entry", None, UNRELATED_SERTONIN_CODE),
        ],
    )
    outcome = _outcome(workbook, client)

    assert _misspelling_codes(outcome) == []


@pytest.mark.req("FR-79")
def test_an_all_uppercase_token_is_never_flagged_by_heuristic_two_either(
    tmp_path: Path,
) -> None:
    """ADR-0007 Decision 5's all-uppercase suspect restriction applies "in
    either heuristic" (this module's own docstring) - not just heuristic 1.
    'ANTENATOL', corpus-rare (1 entry) against a corpus-common 'antenatal'
    (3 entries), is exactly heuristic 2's shape - but rendered in caps, it
    must be silent, the same way an uppercase heuristic-1 suspect is."""
    workbook = _workbook(
        tmp_path,
        [
            ("ANTENATOL", None, None),
            ("antenatal screen", None, None),
            ("antenatal panel", None, None),
            ("antenatal profile", None, None),
        ],
    )
    outcome = _outcome(workbook)

    assert _misspelling_codes(outcome) == []


@pytest.mark.req("FR-79")
@pytest.mark.parametrize(
    ("rare_surface", "rare_message_quoted"),
    [
        ("Gentamic", "'Gentamic'"),  # rare_length = common_length - MAX_EDIT_DISTANCE
        ("Gentamicinnn", "'Gentamicinnn'"),  # rare_length = common_length + MAX_EDIT_DISTANCE
    ],
    ids=["shorter", "longer"],
)
def test_heuristic_two_still_matches_at_both_edges_of_the_length_bucket_window(
    tmp_path: Path, rare_surface: str, rare_message_quoted: str
) -> None:
    """Pins ``_heuristic_two_candidates``'s length-bucketing window at both
    boundaries: 'Gentamic' (8 chars) is ``MAX_EDIT_DISTANCE`` (2) shorter
    than 'Gentamicin' (10 chars), and 'Gentamicinnn' (12 chars, a doubled-
    letter-style typo) is the same distance longer - the near and far edges
    of the ``rare_length +/- MAX_EDIT_DISTANCE`` window this heuristic scans.
    Narrowing the window to only one side (e.g. dropping the lower half,
    ``range(rare_length, rare_length + MAX_EDIT_DISTANCE + 1)``) passes the
    'shorter' case while missing the 'longer' one - both directions need
    their own case. Every token here is at least ``LONG_TOKEN_LENGTH`` (8)
    long, so distance 2 is admissible at all."""
    workbook = _workbook(
        tmp_path,
        [
            ("Gentamicin", None, None),
            ("Gentamicin panel", None, None),
            ("Gentamicin ratio", None, None),
            (rare_surface, None, None),
        ],
    )
    outcome = _outcome(workbook)

    findings = _misspelling_codes(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.INCONSISTENT_SPELLING
    assert rare_message_quoted in findings[0].message
    assert "'Gentamicin'" in findings[0].message


@pytest.mark.req("FR-79")
def test_the_inconsistent_spelling_message_quotes_the_most_frequent_surface(
    tmp_path: Path,
) -> None:
    """'_corpus_index' must pick the representative surface by frequency,
    not alphabetically: an all-uppercase 'ANTENATAL' (1 entry) sorts before
    'Antenatal' (3 entries) lexicographically, but the far more common
    spelling is 'Antenatal', and the finding message must say so. Reverting
    to plain ``min(counts)`` (alphabetical only) passes every other test in
    this suite - none of them assert which surface is quoted."""
    workbook = _workbook(
        tmp_path,
        [
            ("ANTENATAL", None, None),
            ("Antenatal screen", None, None),
            ("Antenatal panel", None, None),
            ("Antenatal profile", None, None),
            ("Antenatel", None, None),  # the rare typo
        ],
    )
    outcome = _outcome(workbook)

    findings = _misspelling_codes(outcome)
    assert len(findings) == 1
    assert findings[0].code == FindingCode.INCONSISTENT_SPELLING
    assert "Antenatel" in findings[0].message
    assert "Antenatal" in findings[0].message
    assert "ANTENATAL" not in findings[0].message


# -- tie-break rule 4: no evidence, silence -----------------------------------


@pytest.mark.req("FR-79")
def test_a_tied_pair_with_no_authority_or_role_evidence_is_silent(tmp_path: Path) -> None:
    """Two near-match nonsense tokens, both in the preferred-term cell of the
    same entry (so rule 3's synonym-vs-preferred-term tie-break does not
    apply either), both absent from any authority set, both with identical
    (zero, since neither appears anywhere else) corpus row-counts: with no
    evidence of which spelling is correct, FR-72 requires a finding to be
    able to state the required action, and there isn't one here."""
    workbook = _workbook(tmp_path, [("Grivetol Grivetal marker", None, None)])
    outcome = _outcome(workbook)

    assert _misspelling_codes(outcome) == []


# -- provenance -----------------------------------------------------------------


@pytest.mark.req("FR-79")
def test_the_run_records_authority_source_and_counts(tmp_path: Path) -> None:
    workbook = _workbook(tmp_path, [("Antenatal screen", "antental", None)])
    outcome = _outcome(workbook)

    assert outcome.run.authority_source is AuthoritySource.WORKBOOK_ONLY
    assert outcome.run.cells_scanned == 2
    assert outcome.run.tokens_considered >= 2
    assert outcome.run.probable_misspelling_count == 1
    assert outcome.run.inconsistent_spelling_count == 0
